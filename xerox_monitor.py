import tkinter as tk
from tkinter import ttk, messagebox
import customtkinter as ctk
import json, os, asyncio, socket, threading, csv, ipaddress, subprocess, sys
import webbrowser
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

PIL_OK = False
try:
    from PIL import Image as PILImage, ImageTk, ImageDraw
    PIL_OK = True
except ImportError:
    pass

TRAY_OK = False
try:
    import pystray
    TRAY_OK = True
except ImportError:
    pass

REQUESTS_OK = False
try:
    import requests as _requests
    from requests.packages.urllib3.exceptions import InsecureRequestWarning
    _requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
    REQUESTS_OK = True
except ImportError:
    pass

_historial_lock = threading.Lock()

# ── SNMP ──────────────────────────────────────────────────────────────────────
SNMP_OK = False
SNMP_ERROR = ""
try:
    from pysnmp.hlapi.v1arch.asyncio import (
        SnmpDispatcher, CommunityData, UdpTransportTarget,
        ObjectType, ObjectIdentity, get_cmd, next_cmd
    )
    SNMP_OK = True
except Exception as e:
    SNMP_ERROR = str(e)

# ── HELPERS CONSUMIBLES ───────────────────────────────────────────────────────
_ABREV_MAP = [
    ("Black Toner",   "Negro"),  ("Cyan Toner",    "Cian"),
    ("Magenta Toner", "Magenta"),("Yellow Toner",  "Amarillo"),
    ("Black Drum",    "Tambor Negro"), ("Cyan Drum","Tambor Cian"),
    ("Magenta Drum",  "Tambor Magenta"),("Yellow Drum","Tambor Amarillo"),
    ("Black",  "Negro"), ("Cyan",  "Cian"),
    ("Magenta","Magenta"),("Yellow","Amarillo"),
    ("Toner",  "Tóner"), ("Drum",  "Tambor"), ("Fuser","Fusor"),
    ("Transfer Roller","Transfer"),("Transfer","Transfer"),
    ("Maintenance","Mant."),("Waste Toner","Residuo"),
    ("Waste",  "Residuo"),("Staple","Grapas"),
    ("Hole Punch","Taladro"),
]

def _abrev_consumible(name):
    n = name
    for orig, rep in _ABREV_MAP:
        if orig.lower() in n.lower():
            n = n.lower().replace(orig.lower(), rep)
            break
    n = n.strip().title()
    return n[:16] if len(n) > 16 else n

def _formato_consumibles(cons, umbral_critico, umbral_alerta):
    if not cons:
        return "—"
    sorted_c = sorted(cons, key=lambda c: c["porcentaje"])
    parts, ok_count = [], 0
    for c in sorted_c:
        pct = c["porcentaje"]
        if pct <= umbral_alerta:
            sym = "⚠" if pct <= umbral_critico else "·"
            parts.append(f"{sym} {_abrev_consumible(c['componente'])}: {pct}%")
        else:
            ok_count += 1
    if ok_count == 1:
        ok_item = next(c for c in cons if c["porcentaje"] > umbral_alerta)
        parts.append(f"✓ {_abrev_consumible(ok_item['componente'])}: {ok_item['porcentaje']}%")
    elif ok_count > 1:
        parts.append(f"✓ {ok_count} en buen nivel")
    return "   ".join(parts) if parts else "✓ Todo OK"

# ── CONSTANTES ────────────────────────────────────────────────────────────────
DB_FILE        = "impresoras.json"
HISTORIAL_FILE = "historial.json"
CONFIG_FILE    = "config.json"
BASE_OID       = "1.3.6.1.2.1.43.11.1.1"
DEFAULT_CONFIG = {
    "umbral_critico": 15,
    "umbral_alerta":  20,
    "comunidad_snmp": "public",
    "timeout_snmp":   2.0,
    "autorefresh_seg": 0,
    "xsa_usuario":    "admin",
    "xsa_password":   "",
    "xsa_autodownload": True,
    "xsa_ultimo_mes": "",   # "YYYY-MM" del último autodownload
}

# ── PALETA ────────────────────────────────────────────────────────────────────
BG      = "#1a1d2e"
BG2     = "#232640"
BG3     = "#2c3057"
ACCENT  = "#4f8ef7"
TEXT    = "#e8eaf0"
TEXT2   = "#8b92b8"
OK      = "#2ecc71"
WARN    = "#f39c12"
CRIT    = "#e74c3c"
OFFLINE = "#6c7a99"
ROW_ALT = "#1e2238"
ROW_SEL = "#2a3560"
BORDER  = "#353860"

# ── PERSISTENCIA ──────────────────────────────────────────────────────────────
def cargar_json(path, default):
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, ValueError):
            # Archivo corrupto: renombrarlo y empezar limpio
            try:
                os.rename(path, path + ".corrupto")
            except OSError:
                pass
    return default

def guardar_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def cargar_impresoras():
    raw = cargar_json(DB_FILE, [])
    out = []
    for item in raw:
        if isinstance(item, str):
            out.append({"ip": item, "nombre": item, "ubicacion": "", "comunidad": "", "imagen": ""})
        else:
            out.append({**item, "imagen": item.get("imagen", "")})
    return out

def guardar_historial(ip, consumibles, nombre=None):
    with _historial_lock:
        h = cargar_json(HISTORIAL_FILE, {})
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if ip not in h:
            h[ip] = []
        entry = {"ts": ts, "consumibles": consumibles}
        if nombre:
            entry["nombre"] = nombre
        h[ip].append(entry)
        h[ip] = h[ip][-100:]
        guardar_json(HISTORIAL_FILE, h)

# ── SNMP ──────────────────────────────────────────────────────────────────────
def limpiar_nombre(n):
    for sep in [";", ", PN", " (", ","]:
        if sep in n:
            n = n.split(sep)[0]
    return n.strip()

async def _snmp_bulk(ip, comunidad, timeout):
    desc, level, maxcap = {}, {}, {}
    dispatcher = SnmpDispatcher()
    try:
        transport = await UdpTransportTarget.create((ip, 161), timeout=timeout, retries=1)
    except Exception as e:
        return None, f"No se pudo conectar: {e}"

    for supply_idx in range(1, 20):
        idx = f"1.{supply_idx}"
        nombre_val = level_val = maxcap_val = None
        for sub, oid_full in [
            (".6.", BASE_OID + ".6." + idx),
            (".8.", BASE_OID + ".8." + idx),
            (".9.", BASE_OID + ".9." + idx),
        ]:
            try:
                errI, errS, _, vb = await get_cmd(
                    dispatcher, CommunityData(comunidad), transport,
                    ObjectType(ObjectIdentity(oid_full))
                )
                if errI or errS:
                    continue
                for oid, val in vb:
                    v = str(val)
                    if "No Such" in v or "End of" in v:
                        continue
                    if sub == ".6.":
                        nombre_val = v
                    elif sub == ".8.":
                        try: maxcap_val = int(val)
                        except: pass
                    elif sub == ".9.":
                        try: level_val = int(val)
                        except: pass
            except Exception:
                pass
        if nombre_val is None:
            break
        desc[idx] = nombre_val
        if maxcap_val is not None: maxcap[idx] = maxcap_val
        if level_val  is not None: level[idx]  = level_val

    resultados = []
    for idx, name in desc.items():
        lvl = level.get(idx)
        maximum = maxcap.get(idx)
        if lvl is None or maximum is None: continue
        if lvl == -2: continue
        if maximum > 0 and lvl >= 0:
            pct = max(0, min(round((lvl / maximum) * 100), 100))
        elif lvl == -3:
            pct = 100
        else:
            continue
        resultados.append({"componente": limpiar_nombre(name), "porcentaje": pct})

    if not resultados:
        return None, "Sin consumibles detectados"
    return resultados, None

async def _snmp_nombre(ip, comunidad, timeout):
    dispatcher = SnmpDispatcher()
    try:
        transport = await UdpTransportTarget.create((ip, 161), timeout=timeout, retries=1)
        errI, errS, _, vb = await get_cmd(
            dispatcher, CommunityData(comunidad), transport,
            ObjectType(ObjectIdentity("1.3.6.1.2.1.1.1.0"))
        )
        if errI or errS: return None
        for _, val in vb:
            n = str(val).strip()
            if n and "No Such" not in n:
                return n.split(";")[0].split(",")[0].strip()[:60]
    except Exception:
        pass
    return None

def obtener_consumibles(ip, comunidad="public", timeout=2.0):
    try:    return asyncio.run(_snmp_bulk(ip, comunidad, timeout))
    except Exception as e: return None, str(e)

def obtener_nombre_snmp(ip, comunidad="public", timeout=2.0):
    try:    return asyncio.run(_snmp_nombre(ip, comunidad, timeout))
    except: return None

# OIDs adicionales útiles
OIDS_INFO = {
    "sys_desc":     "1.3.6.1.2.1.1.1.0",      # Descripción del sistema
    "sys_nombre":   "1.3.6.1.2.1.1.5.0",      # Nombre del host (sysName)
    "sys_uptime":   "1.3.6.1.2.1.1.3.0",      # Uptime del dispositivo
    "paginas":      "1.3.6.1.2.1.43.10.2.1.4.1.1",  # Contador páginas impresas
    "estado_imp":   "1.3.6.1.2.1.25.3.5.1.1.1",     # hrPrinterStatus (idle/printing/error)
    "estado_det":   "1.3.6.1.2.1.25.3.5.1.2.1",     # hrPrinterDetectedErrorState
    "modelo":       "1.3.6.1.2.1.25.3.2.1.3.1",     # hrDeviceDescr (modelo)
    "serial":       "1.3.6.1.2.1.43.5.1.1.17.1",    # prtGeneralSerialNumber
    "bandeja_cap":  "1.3.6.1.2.1.43.8.2.1.9.1.1",   # Capacidad bandeja entrada
    "bandeja_nivel":"1.3.6.1.2.1.43.8.2.1.10.1.1",  # Nivel actual bandeja entrada
}

ESTADO_IMP_MAP = {
    "1": "Otro", "2": "Desconocido", "3": "Idle",
    "4": "Imprimiendo", "5": "Calentando"
}

async def _snmp_info_extra(ip, comunidad, timeout):
    """Obtiene OIDs adicionales: páginas, uptime, estado, serie, bandeja..."""
    dispatcher = SnmpDispatcher()
    info = {}
    try:
        transport = await UdpTransportTarget.create((ip, 161), timeout=timeout, retries=1)
    except Exception as e:
        return info

    for clave, oid in OIDS_INFO.items():
        try:
            errI, errS, _, vb = await get_cmd(
                dispatcher, CommunityData(comunidad), transport,
                ObjectType(ObjectIdentity(oid))
            )
            if errI or errS: continue
            for _, val in vb:
                v = str(val).strip()
                if "No Such" in v or "End of" in v or not v: continue
                info[clave] = v
        except Exception:
            pass
    return info

def obtener_info_extra(ip, comunidad="public", timeout=2.0):
    try:    return asyncio.run(_snmp_info_extra(ip, comunidad, timeout))
    except: return {}

# Mapeo de códigos prtAlertCode → texto en español
_ALERT_CODE_MAP = {
    3:    "Cubierta abierta",
    5:    "Enclavamiento abierto",
    8:    "Atasco de papel",
    16:   "Papel bajo",
    17:   "Sin papel",
    501:  "Consumible bajo",
    502:  "Consumible agotado",
    503:  "Consumible erróneo",
    801:  "Fusor: fin de vida próximo",
    802:  "Fusor: fin de vida",
    901:  "Unidad limpieza: fin próximo",
    902:  "Unidad limpieza: agotada",
    1101: "Pedir suministro ya",
    1102: "Residuos casi llenos",
    1103: "Residuos llenos",
    1104: "Consumible no autenticado",
}
_ALERT_SEV_MAP = {
    "3": ("critical", "🔴"),
    "4": ("warning",  "🟡"),
    "5": ("info",     "🔵"),
}

async def _snmp_alertas_async(ip, comunidad, timeout):
    """Lee la tabla prtAlert (1.3.6.1.2.1.43.18.1.1) y devuelve alertas activas."""
    dispatcher = SnmpDispatcher()
    try:
        transport = await UdpTransportTarget.create((ip, 161), timeout=timeout, retries=1)
    except Exception:
        return []

    alertas = []
    BASE_ALERT = "1.3.6.1.2.1.43.18.1.1"
    # Recorremos índices 1..15 (más que suficiente para cualquier impresora)
    for idx in range(1, 16):
        sev_oid  = f"{BASE_ALERT}.2.1.{idx}"
        code_oid = f"{BASE_ALERT}.8.1.{idx}"
        desc_oid = f"{BASE_ALERT}.11.1.{idx}"
        row = {}
        any_found = False
        for campo, oid in (("sev", sev_oid), ("code", code_oid), ("desc", desc_oid)):
            try:
                errI, errS, _, vb = await get_cmd(
                    dispatcher, CommunityData(comunidad), transport,
                    ObjectType(ObjectIdentity(oid))
                )
                if errI or errS:
                    continue
                for _, val in vb:
                    v = str(val).strip()
                    if "No Such" in v or "End of" in v or not v:
                        continue
                    row[campo] = v
                    any_found = True
            except Exception:
                pass
        if not any_found:
            break  # tabla terminada
        sev = row.get("sev", "0")
        if sev in ("0", "1", "2"):  # 0=desconocido, 1=otro, 2=informacional menor
            continue
        code_raw = row.get("code", "")
        try:
            code_int = int(code_raw)
        except Exception:
            code_int = 0
        desc_raw = row.get("desc", "").strip()
        texto = _ALERT_CODE_MAP.get(code_int) or (desc_raw if desc_raw else f"Código {code_raw}")
        nivel, icono = _ALERT_SEV_MAP.get(sev, ("warning", "🟡"))
        alertas.append({"nivel": nivel, "icono": icono, "texto": texto, "code": code_int})

    # Deduplicar por texto (algunas impresoras repiten la misma alerta)
    seen, unique = set(), []
    for a in alertas:
        if a["texto"] not in seen:
            seen.add(a["texto"])
            unique.append(a)
    return unique

def obtener_alertas(ip, comunidad="public", timeout=2.0):
    try:    return asyncio.run(_snmp_alertas_async(ip, comunidad, timeout))
    except: return []

def _formatear_uptime(raw):
    """Convierte TimeTicks SNMP (centésimas de segundo) a texto legible."""
    try:
        centesimas = int(str(raw).replace(",", "").strip())
        total_seg = centesimas // 100
        dias = total_seg // 86400
        horas = (total_seg % 86400) // 3600
        minutos = (total_seg % 3600) // 60
        if dias:
            return f"{dias}d {horas:02d}h {minutos:02d}m"
        return f"{horas}h {minutos:02d}m"
    except Exception:
        return str(raw)[:30]

# ── ESCANEO DE RED ────────────────────────────────────────────────────────────
def _parsear_rango(texto):
    """Devuelve lista de IPs a escanear desde CIDR, rango A-B o IP única."""
    texto = texto.strip()
    ips = []
    try:
        if "/" in texto:
            red = ipaddress.IPv4Network(texto, strict=False)
            ips = [str(h) for h in red.hosts()]
        elif "-" in texto:
            partes = texto.rsplit("-", 1)
            base = partes[0].rsplit(".", 1)
            prefijo = base[0]
            ini = int(base[1])
            fin = int(partes[1])
            ips = [f"{prefijo}.{i}" for i in range(ini, fin + 1)]
        else:
            ipaddress.IPv4Address(texto)
            ips = [texto]
    except Exception:
        pass
    return ips

async def _snmp_ping_async(ip, comunidad, timeout):
    """Comprueba SNMP e impresora MIB. Solo devuelve resultado si es impresora."""
    dispatcher = SnmpDispatcher()
    try:
        transport = await UdpTransportTarget.create((ip, 161), timeout=timeout, retries=0)
    except Exception:
        return None

    async def get(oid):
        try:
            errI, errS, _, vb = await get_cmd(
                dispatcher, CommunityData(comunidad), transport,
                ObjectType(ObjectIdentity(oid))
            )
            if errI or errS:
                return None
            for _, val in vb:
                v = str(val).strip()
                if "No Such" in v or "End of" in v or not v:
                    return None
                return v
        except Exception:
            return None

    # Verificar Printer MIB: prtMarkerSuppliesDescription (1.3.6.1.2.1.43.11.1.1.6.1.1)
    # Es el OID más fiable para confirmar que es una impresora
    es_impresora = await get("1.3.6.1.2.1.43.11.1.1.6.1.1")
    if not es_impresora:
        # Segunda oportunidad: prtGeneralSerialNumber
        es_impresora = await get("1.3.6.1.2.1.43.5.1.1.17.1")
    if not es_impresora:
        return None

    nombre   = await get("1.3.6.1.2.1.1.5.0") or ""   # sysName
    modelo   = await get("1.3.6.1.2.1.25.3.2.1.3.1") or ""  # hrDeviceDescr
    sys_desc = await get("1.3.6.1.2.1.1.1.0") or ""

    nombre = (nombre or modelo or sys_desc or ip)[:50]
    return {
        "ip": ip,
        "sys_desc": sys_desc[:60],
        "nombre": nombre,
        "modelo": modelo[:40],
    }

def _snmp_ping(ip, comunidad, timeout):
    try:
        return asyncio.run(_snmp_ping_async(ip, comunidad, min(timeout, 1.5)))
    except Exception:
        return None

# ══════════════════════════════════════════════════════════════════════════════
# DIALOGO ESCANEO DE RED
# ══════════════════════════════════════════════════════════════════════════════
class DialogEscaneoRed(ctk.CTkToplevel):
    def __init__(self, parent, cfg, ips_existentes):
        super().__init__(parent)
        self.cfg = cfg
        self.ips_existentes = set(ips_existentes)
        self.resultado = []          # lista de dicts a añadir
        self._cancelar = False
        self._checks = {}            # ip → BooleanVar

        self.title("Escaneo de red")
        self.geometry("620x580")
        self.minsize(540, 480)
        self.configure(fg_color=BG2)
        self.grab_set()
        self.lift()
        self.focus_force()

        # ── Barra inferior fija ──
        btns = ctk.CTkFrame(self, fg_color=BG2, height=56)
        btns.pack(side="bottom", fill="x", padx=20, pady=12)
        btns.pack_propagate(False)
        self.btn_cancel = ctk.CTkButton(
            btns, text="Cancelar", width=130, height=36,
            fg_color=BG3, hover_color=BORDER, text_color=TEXT,
            font=("Segoe UI", 12), command=self._on_cancel)
        self.btn_cancel.pack(side="left", padx=4)
        self.btn_add = ctk.CTkButton(
            btns, text="＋  Añadir seleccionados", width=190, height=36,
            fg_color=ACCENT, hover_color="#3a7de8", text_color=TEXT,
            font=("Segoe UI", 12, "bold"), command=self._añadir_sel,
            state="disabled")
        self.btn_add.pack(side="right", padx=4)

        sep = ctk.CTkFrame(self, fg_color=BORDER, height=1)
        sep.pack(side="bottom", fill="x")

        # ── Cuerpo ──
        body = ctk.CTkFrame(self, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=16, pady=12)

        # Fila de entrada
        row1 = ctk.CTkFrame(body, fg_color="transparent")
        row1.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(row1, text="Rango de red:", text_color=TEXT2,
                     font=("Segoe UI", 11)).pack(side="left")
        self.entry_rango = ctk.CTkEntry(
            row1, placeholder_text="192.168.1.0/24  ó  192.168.1.1-254",
            width=300, height=32, fg_color=BG3, border_color=BORDER, text_color=TEXT,
            font=("Segoe UI", 11))
        self.entry_rango.pack(side="left", padx=8)
        self.btn_scan = ctk.CTkButton(
            row1, text="▶  Escanear", width=110, height=32,
            fg_color=ACCENT, hover_color="#3a7de8", text_color=TEXT,
            font=("Segoe UI", 11, "bold"), command=self._iniciar_scan)
        self.btn_scan.pack(side="left", padx=4)

        # Progreso
        self.lbl_prog = ctk.CTkLabel(body, text="", text_color=TEXT2,
                                      font=("Segoe UI", 10))
        self.lbl_prog.pack(anchor="w")
        self.progress = ctk.CTkProgressBar(body, width=580, height=8,
                                            fg_color=BG3, progress_color=ACCENT)
        self.progress.set(0)
        self.progress.pack(fill="x", pady=(2, 8))

        # Cabecera de resultados
        hdr = ctk.CTkFrame(body, fg_color=BG3, corner_radius=4, height=28)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text="✔", width=30, font=("Segoe UI", 10, "bold"),
                     text_color=TEXT2).pack(side="left", padx=4)
        for txt, w in [("IP", 130), ("Nombre / Hostname", 180), ("Modelo", 200)]:
            ctk.CTkLabel(hdr, text=txt, width=w, font=("Segoe UI", 9, "bold"),
                         text_color=TEXT2, anchor="w").pack(side="left", padx=4)

        # Lista de resultados (scroll)
        self.scroll = ctk.CTkScrollableFrame(body, fg_color="transparent")
        self.scroll.pack(fill="both", expand=True, pady=(2, 0))

        self.lbl_vacio = ctk.CTkLabel(
            self.scroll, text="Introduce un rango y pulsa Escanear",
            text_color=TEXT2, font=("Segoe UI", 11))
        self.lbl_vacio.pack(pady=40)

    def _iniciar_scan(self):
        rango = self.entry_rango.get().strip()
        ips = _parsear_rango(rango)
        if not ips:
            self.lbl_prog.configure(text="⚠  Rango no válido. Ej: 192.168.1.0/24", text_color=CRIT)
            return
        if len(ips) > 1024:
            self.lbl_prog.configure(text="⚠  Rango demasiado grande (máx 1024 IPs)", text_color=CRIT)
            return

        # Limpiar resultados anteriores
        for w in self.scroll.winfo_children():
            w.destroy()
        self._checks.clear()
        self.btn_add.configure(state="disabled")
        self.btn_scan.configure(state="disabled")
        self._cancelar = False
        self.progress.set(0)
        self.lbl_prog.configure(text=f"Escaneando {len(ips)} IPs...", text_color=TEXT2)

        threading.Thread(target=self._scan_worker, args=(ips,), daemon=True).start()

    def _scan_worker(self, ips):
        total = len(ips)
        encontrados = 0
        com = self.cfg["comunidad_snmp"]
        timeout = min(self.cfg["timeout_snmp"], 1.5)

        # Usamos hasta 80 hilos en paralelo — el timeout corto evita bloqueos
        with ThreadPoolExecutor(max_workers=min(80, total)) as ex:
            futures = {ex.submit(_snmp_ping, ip, com, timeout): ip for ip in ips}
            done = 0
            for fut in as_completed(futures):
                if self._cancelar:
                    ex.shutdown(wait=False, cancel_futures=True)
                    break
                done += 1
                resultado = fut.result()
                progreso = done / total
                self.after(0, lambda p=progreso, d=done, t=total:
                           self._actualizar_prog(p, d, t))
                if resultado:
                    encontrados += 1
                    self.after(0, lambda r=resultado: self._agregar_fila(r))

        if not self._cancelar:
            self.after(0, lambda e=encontrados: self._scan_done(e))

    def _actualizar_prog(self, p, done, total):
        self.progress.set(p)
        self.lbl_prog.configure(
            text=f"Escaneando... {done}/{total} IPs  ({int(p*100)}%)",
            text_color=TEXT2)

    def _scan_done(self, encontrados):
        self.btn_scan.configure(state="normal")
        self.progress.set(1)
        if encontrados == 0:
            self.lbl_prog.configure(
                text="Sin dispositivos SNMP encontrados en ese rango.", text_color=WARN)
            self.lbl_vacio = ctk.CTkLabel(
                self.scroll, text="Sin resultados", text_color=TEXT2,
                font=("Segoe UI", 11))
            self.lbl_vacio.pack(pady=40)
        else:
            ya = sum(1 for ip in self._checks if ip in self.ips_existentes)
            nuevos = encontrados - ya
            self.lbl_prog.configure(
                text=f"✔  {encontrados} dispositivos encontrados  ({nuevos} nuevos)",
                text_color=OK)
            if nuevos > 0:
                self.btn_add.configure(state="normal")

    def _agregar_fila(self, r):
        ip = r["ip"]
        ya_existe = ip in self.ips_existentes
        var = tk.BooleanVar(value=not ya_existe)
        self._checks[ip] = (var, r)

        fila = ctk.CTkFrame(self.scroll, fg_color=BG3 if ya_existe else BG2,
                             corner_radius=3, height=30)
        fila.pack(fill="x", pady=1)
        fila.pack_propagate(False)

        cb = ctk.CTkCheckBox(fila, text="", variable=var, width=30,
                              fg_color=ACCENT, hover_color="#3a7de8",
                              state="disabled" if ya_existe else "normal",
                              command=self._check_changed)
        cb.pack(side="left", padx=4)

        color = TEXT2 if ya_existe else TEXT
        ctk.CTkLabel(fila, text=ip, width=130, font=("Consolas", 10),
                     text_color=color, anchor="w").pack(side="left", padx=4)
        nombre = (r["nombre"] or ip)
        sufijo = "  (ya añadida)" if ya_existe else ""
        ctk.CTkLabel(fila, text=nombre + sufijo, width=180,
                     font=("Segoe UI", 10), text_color=color,
                     anchor="w").pack(side="left", padx=4)
        ctk.CTkLabel(fila, text=r.get("modelo","—") or "—", width=200,
                     font=("Segoe UI", 10), text_color=TEXT2,
                     anchor="w").pack(side="left", padx=4)

    def _check_changed(self):
        hay_sel = any(
            var.get() and ip not in self.ips_existentes
            for ip, (var, _) in self._checks.items()
        )
        self.btn_add.configure(state="normal" if hay_sel else "disabled")

    def _añadir_sel(self):
        self.resultado = [
            {"ip": ip, "nombre": r["nombre"] or ip,
             "ubicacion": "", "comunidad": ""}
            for ip, (var, r) in self._checks.items()
            if var.get() and ip not in self.ips_existentes
        ]
        self.destroy()

    def _on_cancel(self):
        self._cancelar = True
        self.destroy()

# ══════════════════════════════════════════════════════════════════════════════
# DIALOGO AÑADIR/EDITAR
# ══════════════════════════════════════════════════════════════════════════════
class DialogImpresora(ctk.CTkToplevel):
    def __init__(self, parent, cfg, imp=None):
        super().__init__(parent)
        self.cfg = cfg
        self.resultado = None
        self.title("Añadir dispositivo" if not imp else "Editar dispositivo")
        self.geometry("420x560")
        self.minsize(380, 520)
        self.resizable(True, True)
        self.configure(fg_color=BG2)
        self.grab_set()
        self.lift()
        self.focus_force()

        # Botones ABAJO fijos
        btns = ctk.CTkFrame(self, fg_color=BG2, height=56)
        btns.pack(side="bottom", fill="x", padx=20, pady=12)
        btns.pack_propagate(False)
        ctk.CTkButton(btns, text="Cancelar", width=130, height=36,
                      fg_color=BG3, hover_color=BORDER, text_color=TEXT,
                      font=("Segoe UI", 12), command=self.destroy).pack(side="left", padx=4)
        ctk.CTkButton(btns, text="✔  Guardar", width=130, height=36,
                      fg_color=ACCENT, hover_color="#3a7de8", text_color=TEXT,
                      font=("Segoe UI", 12, "bold"), command=self._guardar).pack(side="right", padx=4)

        self.status_lbl = ctk.CTkLabel(self, text="", text_color=WARN, font=("Segoe UI", 10))
        self.status_lbl.pack(side="bottom", pady=(0,4))

        sep = ctk.CTkFrame(self, fg_color=BORDER, height=1)
        sep.pack(side="bottom", fill="x")

        # Scroll para el contenido
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=4, pady=4)

        pad = {"padx": 20, "pady": 6}

        ctk.CTkLabel(scroll, text="Dirección IP *", text_color=TEXT2, font=("Segoe UI", 11)).pack(anchor="w", **pad)
        self.ip = ctk.CTkEntry(scroll, placeholder_text="192.168.1.100", width=360, fg_color=BG3, border_color=BORDER, text_color=TEXT)
        self.ip.pack(**pad)
        if imp: self.ip.insert(0, imp.get("ip", ""))

        ctk.CTkLabel(scroll, text="Nombre (vacío = autodetectar por SNMP)", text_color=TEXT2, font=("Segoe UI", 11)).pack(anchor="w", **pad)
        self.nombre = ctk.CTkEntry(scroll, placeholder_text="Xerox Planta 2", width=360, fg_color=BG3, border_color=BORDER, text_color=TEXT)
        self.nombre.pack(**pad)
        if imp: self.nombre.insert(0, imp.get("nombre", ""))

        ctk.CTkLabel(scroll, text="Ubicación", text_color=TEXT2, font=("Segoe UI", 11)).pack(anchor="w", **pad)
        self.ubic = ctk.CTkEntry(scroll, placeholder_text="Sala de reuniones A", width=360, fg_color=BG3, border_color=BORDER, text_color=TEXT)
        self.ubic.pack(**pad)
        if imp: self.ubic.insert(0, imp.get("ubicacion", ""))

        ctk.CTkLabel(scroll, text="Comunidad SNMP", text_color=TEXT2, font=("Segoe UI", 11)).pack(anchor="w", **pad)
        self.com = ctk.CTkEntry(scroll, placeholder_text="public", width=360, fg_color=BG3, border_color=BORDER, text_color=TEXT)
        self.com.pack(**pad)
        if imp: self.com.insert(0, imp.get("comunidad", ""))

        ctk.CTkLabel(scroll, text="Imagen del dispositivo (opcional)", text_color=TEXT2, font=("Segoe UI", 11)).pack(anchor="w", **pad)
        img_row = ctk.CTkFrame(scroll, fg_color="transparent")
        img_row.pack(fill="x", **pad)
        self.img_path = ctk.CTkEntry(img_row, placeholder_text="Ruta a la imagen...", width=280,
                                      fg_color=BG3, border_color=BORDER, text_color=TEXT,
                                      state="readonly")
        self.img_path.pack(side="left")
        ctk.CTkButton(img_row, text="Examinar...", width=80, height=28,
                      fg_color=BG3, hover_color=BORDER, text_color=TEXT,
                      font=("Segoe UI", 11), command=self._examinar_imagen).pack(side="left", padx=(6, 0))
        if imp and imp.get("imagen"):
            self.img_path.configure(state="normal")
            self.img_path.insert(0, imp.get("imagen", ""))
            self.img_path.configure(state="readonly")

    def _examinar_imagen(self):
        from tkinter.filedialog import askopenfilename
        path = askopenfilename(
            title="Seleccionar imagen",
            filetypes=[("Imágenes", "*.png *.jpg *.jpeg *.bmp"), ("Todos", "*.*")])
        if path:
            self.img_path.configure(state="normal")
            self.img_path.delete(0, "end")
            self.img_path.insert(0, path)
            self.img_path.configure(state="readonly")

    def _guardar(self):
        ip = self.ip.get().strip()
        if not ip:
            self.status_lbl.configure(text="⚠  La IP es obligatoria.")
            return
        nombre = self.nombre.get().strip()
        comunidad = self.com.get().strip() or self.cfg["comunidad_snmp"]
        if not nombre:
            self.status_lbl.configure(text="⏳  Autodetectando nombre...")
            self.update()
            nombre = obtener_nombre_snmp(ip, comunidad, self.cfg["timeout_snmp"]) or ip
        self.resultado = {
            "ip": ip, "nombre": nombre,
            "ubicacion": self.ubic.get().strip(),
            "comunidad": self.com.get().strip(),
            "imagen": self.img_path.get().strip()
        }
        self.destroy()

# ══════════════════════════════════════════════════════════════════════════════
# DIALOGO CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════════════════════
class DialogConfig(ctk.CTkToplevel):
    def __init__(self, parent, cfg, on_save):
        super().__init__(parent)
        self.cfg = cfg
        self.on_save = on_save
        self.title("Configuración")
        self.geometry("420x560")
        self.minsize(380, 520)
        self.resizable(True, True)
        self.configure(fg_color=BG2)
        self.grab_set()
        self.lift()
        self.focus_force()

        # Botones ABAJO fijos (antes del scroll)
        btns = ctk.CTkFrame(self, fg_color=BG2, height=56)
        btns.pack(side="bottom", fill="x", padx=20, pady=12)
        btns.pack_propagate(False)
        ctk.CTkButton(btns, text="Cancelar", width=130, height=36,
                      fg_color=BG3, hover_color=BORDER, text_color=TEXT,
                      font=("Segoe UI", 12), command=self.destroy).pack(side="left", padx=4)
        ctk.CTkButton(btns, text="✔  Guardar", width=130, height=36,
                      fg_color=ACCENT, hover_color="#3a7de8", text_color=TEXT,
                      font=("Segoe UI", 12, "bold"), command=self._guardar).pack(side="right", padx=4)

        sep = ctk.CTkFrame(self, fg_color=BORDER, height=1)
        sep.pack(side="bottom", fill="x")

        # Scroll para el contenido
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=4, pady=4)

        pad = {"padx": 20, "pady": 5}

        def lbl(text):
            ctk.CTkLabel(scroll, text=text, text_color=TEXT2, font=("Segoe UI", 11)).pack(anchor="w", **pad)

        lbl("Umbral crítico (%)")
        self.crit_var = tk.IntVar(value=cfg["umbral_critico"])
        self.crit_lbl = ctk.CTkLabel(scroll, text=str(self.crit_var.get()), text_color=CRIT, font=("Segoe UI", 13, "bold"))
        self.crit_lbl.pack(anchor="w", padx=24)
        ctk.CTkSlider(scroll, from_=1, to=50, variable=self.crit_var, width=340,
                      command=lambda v: self.crit_lbl.configure(text=str(int(v)))).pack(**pad)

        lbl("Umbral alerta (%)")
        self.alert_var = tk.IntVar(value=cfg["umbral_alerta"])
        self.alert_lbl = ctk.CTkLabel(scroll, text=str(self.alert_var.get()), text_color=WARN, font=("Segoe UI", 13, "bold"))
        self.alert_lbl.pack(anchor="w", padx=24)
        ctk.CTkSlider(scroll, from_=1, to=70, variable=self.alert_var, width=340,
                      command=lambda v: self.alert_lbl.configure(text=str(int(v)))).pack(**pad)

        lbl("Comunidad SNMP")
        self.com = ctk.CTkEntry(scroll, width=340, fg_color=BG3, border_color=BORDER, text_color=TEXT)
        self.com.insert(0, cfg["comunidad_snmp"])
        self.com.pack(**pad)

        lbl("Timeout SNMP (segundos)")
        self.timeout = ctk.CTkEntry(scroll, width=340, fg_color=BG3, border_color=BORDER, text_color=TEXT)
        self.timeout.insert(0, str(cfg["timeout_snmp"]))
        self.timeout.pack(**pad)

        lbl("Auto-refresco")
        self.autoref = ctk.CTkOptionMenu(scroll, values=["Desactivado","30s","60s","2 min","5 min"],
                                          width=340, fg_color=BG3, button_color=ACCENT, text_color=TEXT)
        mapa = {0:"Desactivado", 30:"30s", 60:"60s", 120:"2 min", 300:"5 min"}
        self.autoref.set(mapa.get(cfg["autorefresh_seg"], "Desactivado"))
        self.autoref.pack(**pad)

        ctk.CTkFrame(scroll, fg_color=BORDER, height=1).pack(fill="x", padx=20, pady=(14,4))
        ctk.CTkLabel(scroll, text="Descarga automática XSA (día 1 de cada mes)",
                     text_color=TEXT2, font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=20, pady=(4,0))

        lbl("Usuario web impresora")
        self.xsa_user = ctk.CTkEntry(scroll, width=340, fg_color=BG3, border_color=BORDER, text_color=TEXT)
        self.xsa_user.insert(0, cfg.get("xsa_usuario", "admin"))
        self.xsa_user.pack(**pad)

        lbl("Contraseña web impresora")
        self.xsa_pass = ctk.CTkEntry(scroll, width=340, fg_color=BG3, border_color=BORDER,
                                      text_color=TEXT, show="•")
        self.xsa_pass.insert(0, cfg.get("xsa_password", ""))
        self.xsa_pass.pack(**pad)

        self.xsa_auto_var = tk.BooleanVar(value=cfg.get("xsa_autodownload", True))
        ctk.CTkCheckBox(scroll, text="Activar descarga automática el día 1",
                        variable=self.xsa_auto_var, text_color=TEXT,
                        font=("Segoe UI", 11), fg_color=ACCENT).pack(anchor="w", padx=24, pady=6)

    def _guardar(self):
        try:
            mapa_inv = {"Desactivado":0,"30s":30,"60s":60,"2 min":120,"5 min":300}
            self.cfg.update({
                "umbral_critico":    int(self.crit_var.get()),
                "umbral_alerta":     int(self.alert_var.get()),
                "comunidad_snmp":    self.com.get().strip() or "public",
                "timeout_snmp":      float(self.timeout.get()),
                "autorefresh_seg":   mapa_inv.get(self.autoref.get(), 0),
                "xsa_usuario":       self.xsa_user.get().strip() or "admin",
                "xsa_password":      self.xsa_pass.get(),
                "xsa_autodownload":  self.xsa_auto_var.get(),
            })
            guardar_json(CONFIG_FILE, self.cfg)
            self.on_save()
            self.destroy()
        except ValueError:
            pass

# ══════════════════════════════════════════════════════════════════════════════
# VENTANA PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# DIALOGO CONTABILIDAD POR USUARIO (XSA CSV)
# ══════════════════════════════════════════════════════════════════════════════
CONTABILIDAD_FILE   = "contabilidad_xsa.json"
XSA_PATH_GENERATE   = "/properties/accounting/XSA_generate_date.php"
XSA_PATH_DOWNLOAD   = "/properties/accounting/download_csv.php"

def _xsa_descargar_csv(ip, password, usuario="admin", timeout=12):
    """Descarga el CSV XSA desde la web de la impresora. Devuelve (texto_csv, error)."""
    if not REQUESTS_OK:
        return None, "Instala requests:  pip install requests"
    import re as _re, time as _time
    try:
        base  = f"https://{ip}"
        redir = "/properties/accounting/usageReport.php?from=Acct_Home"
        login = f"/properties/authentication/login.php?redir={redir}"

        s = _requests.Session()
        s.verify  = False
        s.timeout = timeout
        s.headers.update({
            "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "es-ES,es;q=0.9,en-US;q=0.8,en;q=0.7",
        })

        # 1. Login
        r = s.get(base + login, allow_redirects=True)
        csrf = (_re.search(r'name=["\']CSRFToken["\'][^>]*value=["\']([^"\']+)["\']', r.text) or
                _re.search(r'value=["\']([^"\']+)["\'][^>]*name=["\']CSRFToken["\']', r.text))
        s.post(base + "/userpost/xerox.set", allow_redirects=True,
               headers={"Referer": base + login, "Origin": base},
               data={"_fun_function": "HTTP_Authenticate_fn",
                     "NextPage":      "/properties/authentication/luidLogin.php?type=&authStatus=",
                     "frmwebUsername": usuario, "frmwebPassword": password,
                     "frmaltDomain":  "default",
                     "CSRFToken":     csrf.group(1) if csrf else ""})

        # 2. Cargar página de contabilidad para obtener CSRF fresco
        s.cookies.set("scnMboxSelected", "n1", domain=ip)
        s.cookies.set("scnMboxNumNodes", "8",  domain=ip)
        s.cookies.set("propSelected",    "n2",  domain=ip)
        s.cookies.set("propHierarchy",   "00000001010000000000000000000000", domain=ip)
        r = s.get(base + redir, allow_redirects=True,
                  headers={"Referer": base + "/properties/dataManagement/autoConfiguration.php"})
        if "login.php" in r.url:
            return None, "Autenticación fallida — comprueba usuario y contraseña"

        # 3. POST xerox.set para generar el CSV en el servidor
        csrf2 = (_re.search(r'name=["\']CSRFToken["\'][^>]*value=["\']([^"\']+)["\']', r.text) or
                 _re.search(r'value=["\']([^"\']+)["\'][^>]*name=["\']CSRFToken["\']', r.text))
        ts = int(_time.time() * 1000)
        s.post(f"{base}/userpost/xerox.set?ajts{ts}",
               headers={
                   "Referer":          base + redir,
                   "Origin":           base,
                   "X-Requested-With": "XMLHttpRequest",
                   "Accept":           "application/json, text/javascript, */*; q=0.01",
                   "Content-Type":     "application/x-www-form-urlencoded; charset=UTF-8",
               },
               data={"_fun_function": "HTTP_Generate_XSA_Usage_Report_fn",
                     "CSRFToken":     csrf2.group(1) if csrf2 else "",
                     "ShowUserId":    "TRUE",
                     "isAjax":        "true"})

        # 4. Trigger de redirect + descarga
        _time.sleep(0.5)
        s.get(base + XSA_PATH_GENERATE, allow_redirects=False,
              headers={"Referer": base + redir})
        r = s.get(base + XSA_PATH_DOWNLOAD, allow_redirects=False,
                  headers={"Referer": base + redir,
                           "Accept":  "text/csv,application/octet-stream,*/*;q=0.8"})
        if r.status_code == 200 and len(r.content) > 100:
            ct = r.headers.get("Content-Type", "")
            cd = r.headers.get("Content-Disposition", "")
            if ("csv" in ct or "force-download" in ct or "octet" in ct or cd or
                    not r.text.strip().startswith("<")):
                return r.text, None
        return None, f"Respuesta inesperada del servidor ({r.status_code}, {len(r.content)} bytes)"
    except Exception as e:
        return None, str(e)

_MESES_ES = {
    "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
    "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
    "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre",
}

def _mes_label(mes_key):
    """Convierte '2026-05' → 'Mayo 2026'."""
    try:
        y, m = mes_key.split("-")
        return f"{_MESES_ES.get(m, m)} {y}"
    except Exception:
        return mes_key


class DialogContabilidad(ctk.CTkToplevel):
    """Versión ventana flotante (mantenida por compatibilidad)."""
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Contabilidad por usuario")
        self.geometry("1100x680")
        self.minsize(800, 520)
        self.configure(fg_color=BG2)
        self.grab_set()
        self.lift()
        self.focus_force()

        self._datos = cargar_json(CONTABILIDAD_FILE, {})
        self._migrar_formato_antiguo()
        self._sel_impresora = "Todas"
        self._sel_mes = "Acumulado"
        self._sel_vista = "Acumulado"
        self._sort_col = None
        self._sort_rev = False
        self._collapsed = {"DEPARTAMENTOS": False, "USUARIOS": False}
        self._build()


class DialogContabilidadEmbebida(ctk.CTkFrame):
    """Versión embebida en el panel principal (sin ventana flotante)."""
    def __init__(self, parent, app):
        super().__init__(parent, fg_color=BG2, corner_radius=0)
        self._app = app

        self._datos = cargar_json(CONTABILIDAD_FILE, {})
        self._migrar_formato_antiguo()
        self._sel_impresora = "Todas"
        self._sel_mes = "Acumulado"
        self._sel_vista = "Acumulado"
        self._sort_col = None
        self._sort_rev = False
        self._collapsed = {"DEPARTAMENTOS": False, "USUARIOS": False}
        self._build()

    def _migrar_formato_antiguo(self):
        """Convierte datos del formato viejo {ip:{ts,usuarios}} al nuevo {ip:{snapshots:{}}}."""
        cambiado = False
        for ip, bloque in self._datos.items():
            if "usuarios" in bloque and "snapshots" not in bloque:
                ts  = bloque.get("ts", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                mes = ts[:7] if ts else datetime.now().strftime("%Y-%m")
                self._datos[ip] = {
                    "nombre_impresora": bloque.get("nombre_impresora", ip),
                    "snapshots": {mes: {"ts": ts, "usuarios": bloque["usuarios"]}},
                }
                cambiado = True
        if cambiado:
            guardar_json(CONTABILIDAD_FILE, self._datos)

    def _build(self):
        # ── Toolbar ──
        tb = ctk.CTkFrame(self, fg_color=BG3, height=44, corner_radius=0)
        tb.pack(fill="x")
        tb.pack_propagate(False)

        ctk.CTkLabel(tb, text="Contabilidad por usuario — Xerox Standard Accounting",
                     font=("Segoe UI", 12, "bold"), text_color=TEXT).pack(side="left", padx=14)
        ctk.CTkButton(tb, text="🗑  Limpiar", width=90, height=30,
                      fg_color=BG2, hover_color=BORDER, text_color=TEXT2,
                      font=("Segoe UI", 11),
                      command=self._limpiar).pack(side="right", padx=4, pady=7)
        ctk.CTkButton(tb, text="↑  Exportar Excel", width=145, height=30,
                      fg_color=BG3, hover_color=BORDER, text_color=TEXT,
                      font=("Segoe UI", 11),
                      command=self._exportar_excel).pack(side="right", padx=2, pady=7)
        ctk.CTkButton(tb, text="↓  Importar CSV", width=130, height=30,
                      fg_color=BG3, hover_color=BORDER, text_color=TEXT,
                      font=("Segoe UI", 11),
                      command=self._importar_csv).pack(side="right", padx=2, pady=7)
        ctk.CTkButton(tb, text="⟳  Descargar de impresoras", width=190, height=30,
                      fg_color=ACCENT, hover_color="#3a7de8", text_color=TEXT,
                      font=("Segoe UI", 11, "bold"),
                      command=self._descargar_auto).pack(side="right", padx=4, pady=7)

        # ── Fila 1: Impresora ──
        row1 = ctk.CTkFrame(self, fg_color=BG2, height=38)
        row1.pack(fill="x", padx=10, pady=(6, 0))
        row1.pack_propagate(False)
        ctk.CTkLabel(row1, text="Impresora:", text_color=TEXT2,
                     font=("Segoe UI", 11)).pack(side="left", padx=(4, 6))
        self._imp_seg = ctk.CTkSegmentedButton(
            row1, values=self._imp_values(),
            command=self._on_imp_change,
            font=("Segoe UI", 10),
            fg_color=BG3, selected_color=ACCENT, selected_hover_color="#3a7de8",
            unselected_color=BG3, unselected_hover_color=BORDER, text_color=TEXT)
        self._imp_seg.set("Todas")
        self._imp_seg.pack(side="left", padx=4)

        # ── Fila 2: Mes + Vista ──
        row2 = ctk.CTkFrame(self, fg_color=BG2, height=38)
        row2.pack(fill="x", padx=10, pady=(4, 2))
        row2.pack_propagate(False)
        ctk.CTkLabel(row2, text="Mes:", text_color=TEXT2,
                     font=("Segoe UI", 11)).pack(side="left", padx=(4, 6))
        self._mes_combo = ctk.CTkComboBox(
            row2, values=self._mes_values(),
            command=self._on_mes_change,
            width=160, height=28, fg_color=BG3, border_color=BORDER,
            button_color=ACCENT, button_hover_color="#3a7de8",
            dropdown_fg_color=BG3, text_color=TEXT, font=("Segoe UI", 10))
        mes_vals = self._mes_values()
        self._mes_combo.set(mes_vals[0] if mes_vals else "Acumulado")
        self._sel_mes = mes_vals[0] if mes_vals else "Acumulado"
        self._mes_combo.pack(side="left", padx=4)

        ctk.CTkLabel(row2, text="Vista:", text_color=TEXT2,
                     font=("Segoe UI", 11)).pack(side="left", padx=(16, 6))
        self._vista_seg = ctk.CTkSegmentedButton(
            row2, values=["Mensual", "Acumulado"],
            command=self._on_vista_change,
            font=("Segoe UI", 10),
            fg_color=BG3, selected_color=ACCENT, selected_hover_color="#3a7de8",
            unselected_color=BG3, unselected_hover_color=BORDER, text_color=TEXT)
        self._vista_seg.set("Acumulado")
        self._vista_seg.pack(side="left", padx=4)

        # ── Tabla ──
        tbl_frame = ctk.CTkFrame(self, fg_color=BG2, corner_radius=0)
        tbl_frame.pack(fill="both", expand=True, padx=8, pady=(4, 0))

        cols = ("impresora", "usuario", "imp_bw", "imp_color",
                "cop_bw", "cop_color", "total")
        self.tree = ttk.Treeview(tbl_frame, columns=cols,
                                  show="headings", style="F.Treeview", selectmode="browse")
        for col, hdr, w, stretch in [
            ("impresora", "Impresora",      160, True),
            ("usuario",   "Usuario",        150, True),
            ("imp_bw",    "Imp. B/N",        90, False),
            ("imp_color", "Imp. Color",      90, False),
            ("cop_bw",    "Cop. B/N",        90, False),
            ("cop_color", "Cop. Color",      90, False),
            ("total",     "Total páginas",  110, False),
        ]:
            self.tree.heading(col, text=hdr, command=lambda c=col: self._sort(c))
            self.tree.column(col, width=w,
                             anchor="center" if col not in ("impresora", "usuario") else "w",
                             stretch=stretch)

        vsb = ttk.Scrollbar(tbl_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

        self.tree.tag_configure("par",       background=BG2,      foreground=TEXT)
        self.tree.tag_configure("impar",     background=ROW_ALT,  foreground=TEXT)
        self.tree.tag_configure("top",       background="#1a2e1a", foreground=OK)
        self.tree.tag_configure("totales",   background=BG3,      foreground=ACCENT)
        self.tree.tag_configure("seccion",   background="#1c2a3a", foreground="#7ab4f5",
                                font=("Segoe UI", 10, "bold"))
        self.tree.bind("<ButtonRelease-1>", self._on_tree_click)

        # ── Status bar ──
        self.lbl_info = ctk.CTkLabel(self, text="", font=("Segoe UI", 10), text_color=TEXT2)
        self.lbl_info.pack(anchor="w", padx=14, pady=(2, 6))

        self._poblar()

    # ── helpers for filter controls ──────────────────────────────────────────
    def _imp_values(self):
        vals = ["Todas"]
        for ip, bloque in self._datos.items():
            nombre = bloque.get("nombre_impresora", ip)
            vals.append(nombre)
        return vals

    def _mes_values(self):
        """Return month list most-recent first, with 'Acumulado' at top."""
        meses = set()
        ip_filter = self._sel_impresora
        for ip, bloque in self._datos.items():
            nombre = bloque.get("nombre_impresora", ip)
            if ip_filter != "Todas" and nombre != ip_filter:
                continue
            for mk in bloque.get("snapshots", {}).keys():
                meses.add(mk)
        sorted_meses = sorted(meses, reverse=True)
        return ["Acumulado"] + [_mes_label(mk) for mk in sorted_meses]

    def _mes_key_from_label(self, label):
        """Reverse _mes_label: 'Mayo 2026' → '2026-05'. Returns None for 'Acumulado'."""
        if label == "Acumulado":
            return None
        meses_inv = {v: k for k, v in _MESES_ES.items()}
        parts = label.rsplit(" ", 1)
        if len(parts) == 2:
            nombre_mes, anio = parts
            m = meses_inv.get(nombre_mes)
            if m:
                return f"{anio}-{m}"
        return None

    def _on_imp_change(self, value):
        self._sel_impresora = value
        # Refresh mes combo
        new_mes_vals = self._mes_values()
        self._mes_combo.configure(values=new_mes_vals)
        if self._sel_mes not in new_mes_vals:
            self._sel_mes = new_mes_vals[0] if new_mes_vals else "Acumulado"
            self._mes_combo.set(self._sel_mes)
        self._poblar()

    def _on_mes_change(self, value):
        self._sel_mes = value
        self._poblar()

    def _on_vista_change(self, value):
        self._sel_vista = value
        self._poblar()

    def _update_imp_col_visibility(self, has_color=True):
        if self._sel_impresora == "Todas":
            self.tree.column("impresora", width=160, stretch=True)
            self.tree.heading("impresora", text="Impresora")
        else:
            self.tree.column("impresora", width=0, stretch=False, minwidth=0)
            self.tree.heading("impresora", text="")
        # Columnas color: ocultar si no hay ningún valor
        if has_color:
            self.tree.column("imp_color", width=90, stretch=False, minwidth=40)
            self.tree.heading("imp_color", text="Imp. Color")
            self.tree.column("cop_color", width=90, stretch=False, minwidth=40)
            self.tree.heading("cop_color", text="Cop. Color")
        else:
            for col in ("imp_color", "cop_color"):
                self.tree.column(col, width=0, stretch=False, minwidth=0)
                self.tree.heading(col, text="")

    def _descargar_auto(self):
        if not REQUESTS_OK:
            messagebox.showerror("Falta dependencia",
                "Instala requests para usar la descarga automática:\n\n  pip install requests")
            return

        try:
            impresoras = getattr(self, '_app', self.master).impresoras
        except Exception:
            impresoras = []

        if not impresoras:
            messagebox.showinfo("", "No hay impresoras en la lista.")
            return

        dlg = _DialogCredencialesXSA(self, impresoras)
        self.wait_window(dlg)
        if not dlg.resultado:
            return

        seleccionadas = dlg.resultado
        self.lbl_info.configure(text="⏳  Descargando informes...", text_color=WARN)
        self.update()

        def worker():
            errores = []
            ok = 0
            for imp in seleccionadas:
                csv_txt, err = _xsa_descargar_csv(
                    imp["ip"], imp["password"], imp.get("usuario", "admin"))
                if err:
                    errores.append(f"{imp['nombre']}: {err}")
                    continue
                try:
                    filas = self._parsear_csv_texto(csv_txt)
                    if filas:
                        mes_key = datetime.now().strftime("%Y-%m")
                        if imp["ip"] not in self._datos:
                            self._datos[imp["ip"]] = {
                                "nombre_impresora": f"{imp['nombre']} ({imp['ip']})",
                                "snapshots": {}
                            }
                        self._datos[imp["ip"]]["nombre_impresora"] = f"{imp['nombre']} ({imp['ip']})"
                        self._datos[imp["ip"]]["snapshots"][mes_key] = {
                            "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "usuarios": filas,
                        }
                        ok += 1
                    else:
                        errores.append(f"{imp['nombre']}: CSV sin datos de usuario")
                except Exception as e:
                    errores.append(f"{imp['nombre']}: {e}")

            guardar_json(CONTABILIDAD_FILE, self._datos)

            def done():
                self._refresh_controls()
                self._poblar()
                if errores:
                    messagebox.showwarning("Descarga parcial",
                        f"✔ {ok} impresora(s) descargadas.\n\nErrores:\n" + "\n".join(errores))
                else:
                    messagebox.showinfo("Descarga completada",
                        f"✔ {ok} impresora(s) actualizadas correctamente.")
            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _refresh_controls(self):
        """Rebuild the impresora segmented button and mes combo after data changes."""
        self._imp_seg.configure(values=self._imp_values())
        if self._sel_impresora not in self._imp_values():
            self._sel_impresora = "Todas"
            self._imp_seg.set("Todas")
        new_mes_vals = self._mes_values()
        self._mes_combo.configure(values=new_mes_vals)
        if self._sel_mes not in new_mes_vals:
            self._sel_mes = new_mes_vals[0] if new_mes_vals else "Acumulado"
            self._mes_combo.set(self._sel_mes)

    def _parsear_csv_texto(self, texto):
        import io
        for sep in (";", ",", "\t"):
            reader = csv.DictReader(io.StringIO(texto), delimiter=sep)
            try:
                filas = list(reader)
                if filas and len(filas[0]) > 2:
                    return self._procesar_filas(filas)
            except Exception:
                pass
        return []

    def _importar_csv(self):
        from tkinter.filedialog import askopenfilename
        path = askopenfilename(
            title="Selecciona el CSV de Contabilidad Estándar Xerox",
            filetypes=[("CSV", "*.csv"), ("Todos", "*.*")])
        if not path:
            return

        ips = []
        try:
            parent_app = getattr(self, '_app', self.master)
            ips = [f"{i['nombre']} ({i['ip']})" for i in parent_app.impresoras]
        except Exception:
            pass

        dlg = _DialogSeleccionarImpresora(self, ips)
        self.wait_window(dlg)
        if not dlg.resultado:
            return
        ip_label = dlg.resultado
        import re
        m = re.search(r'\(([^)]+)\)', ip_label)
        ip_key = m.group(1) if m else ip_label

        try:
            filas = self._parsear_csv(path)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el CSV:\n{e}")
            return

        if not filas:
            messagebox.showwarning("CSV vacío", "No se encontraron datos de usuario en el archivo.")
            return

        mes_key = datetime.now().strftime("%Y-%m")
        if ip_key not in self._datos:
            self._datos[ip_key] = {"nombre_impresora": ip_label, "snapshots": {}}
        self._datos[ip_key]["nombre_impresora"] = ip_label
        self._datos[ip_key]["snapshots"][mes_key] = {
            "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "usuarios": filas,
        }
        guardar_json(CONTABILIDAD_FILE, self._datos)
        self._refresh_controls()
        self._poblar()
        messagebox.showinfo("Importado", f"✔  {len(filas)} usuarios importados de:\n{os.path.basename(path)}")

    _MAP_COLS = {
        # Xerox AltaLink español (columnas reales del CSV exportado)
        "nombre de cuenta":    "usuario",
        "id de cuenta":        "id",
        "tipo de cuenta":      "tipo",
        # Impresiones negro
        "contador: total de impresiones impresas y copias negro": "imp_bw",
        "contador: total de impresiones negro":                   "imp_bw",
        "impresiones en b/n":  "imp_bw", "impresiones b/n": "imp_bw",
        # Impresiones color
        "contador: total de impresiones impresas y copias color": "imp_color",
        "contador: total de impresiones color":                   "imp_color",
        "impresiones en color": "imp_color", "impresiones color": "imp_color",
        # Copias negro
        "contador: total de impresiones copiadas en negro": "cop_bw",
        "copias en b/n": "cop_bw", "copias b/n": "cop_bw",
        # Copias color
        "contador: total de impresiones copiadas en color": "cop_color",
        "copias en color": "cop_color", "copias color": "cop_color",
        # Escáner
        "contador: total de imágenes escaneadas": "scan",
        "contador: total de escaneados":          "scan",
        "escaneados": "scan", "escáner": "scan", "escaneos": "scan",
        "total": "total",
        # Inglés
        "nombre de usuario": "usuario", "nombre usuario": "usuario",
        "usuario": "usuario", "id de usuario": "id", "id usuario": "id",
        "user name": "usuario", "username": "usuario", "account name": "usuario",
        "user id": "id", "account id": "id",
        "black and white prints": "imp_bw", "b&w prints": "imp_bw",
        "color prints": "imp_color",
        "black and white copies": "cop_bw", "b&w copies": "cop_bw",
        "color copies": "cop_color",
        "scans": "scan", "scan images": "scan",
        "fax images": "fax",
    }

    _DEPARTAMENTOS = {d.lower() for d in (
        "ADMINISTRACIÓN", "DIRECCIÓN", "VICEDIRECCIÓN", "SECRETARÍA",
        "XEFATURA DE ESTUDOS", "RECURSOS", "ALEMÁN", "PORTUGUÉS",
        "FRANCÉS", "ITALIANO", "GALEGO", "PLAMBE_EDLG", "INGLÉS", "SUSTITUTO",
    )}

    _EXCLUIR_USUARIOS = {u.lower() for u in (
        "System User", "CUENTA GENERAL", "Customer Service Engineer Account",
        "Xerox Administrative Group", "Admin", "Diagnostics", "Local System User",
        "Print Exceptions Group", "10.55.161.196", "Guest",
        "IPP Exception Group", "IPP Exception User",
    )}

    def _procesar_filas(self, filas_raw):
        def norm(k): return self._MAP_COLS.get(k.lower().strip(), k.lower().strip())
        resultado = []
        for fila in filas_raw:
            row = {norm(k): (v or "").strip() for k, v in fila.items()}
            nombre = row.get("usuario", "")
            # Ignorar filas de totales, vacías o usuarios del sistema
            if (not nombre
                    or row.get("tipo", "").lower() in ("total", "totales", "system")
                    or nombre.lower() in self._EXCLUIR_USUARIOS):
                continue
            def safe_int(k):
                try: return int(row.get(k, "0").replace(",","").replace(".","") or 0)
                except: return 0
            imp_bw = safe_int("imp_bw"); imp_color = safe_int("imp_color")
            cop_bw = safe_int("cop_bw"); cop_color = safe_int("cop_color")
            scan   = safe_int("scan")
            resultado.append({
                "usuario": row["usuario"], "id": row.get("id",""),
                "imp_bw": imp_bw, "imp_color": imp_color,
                "cop_bw": cop_bw, "cop_color": cop_color,
                "scan": scan, "total": imp_bw+imp_color+cop_bw+cop_color,
            })
        resultado.sort(key=lambda r: r["total"], reverse=True)
        return resultado

    def _parsear_csv(self, path):
        import io
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            contenido = f.read()
        return self._parsear_csv_texto(contenido)

    def _poblar(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        if not self._datos:
            self.lbl_info.configure(
                text="Sin datos. Usa Herramientas → Contabilidad por usuario → Importar CSV.",
                text_color=TEXT2)
            return

        # Determine which IPs to show
        ip_filter = self._sel_impresora
        mes_label_sel = self._sel_mes
        mes_key_sel = self._mes_key_from_label(mes_label_sel)  # None = Acumulado
        vista = self._sel_vista

        # Build rows
        all_rows = []  # list of (nombre_imp, usuario, imp_bw, imp_color, cop_bw, cop_color, scan, total, nota)
        ts_mostrar = []

        for ip_key, bloque in self._datos.items():
            nombre_imp = bloque.get("nombre_impresora", ip_key)
            if ip_filter != "Todas" and nombre_imp != ip_filter:
                continue

            snapshots = bloque.get("snapshots", {})

            # Determine current snapshot key
            if mes_key_sel is not None:
                cur_key = mes_key_sel
            else:
                # Acumulado: use latest snapshot
                sorted_keys = sorted(snapshots.keys(), reverse=True)
                cur_key = sorted_keys[0] if sorted_keys else None

            if cur_key is None or cur_key not in snapshots:
                continue

            cur_snap = snapshots[cur_key]
            cur_usuarios = {u["usuario"]: u for u in cur_snap.get("usuarios", [])}
            ts_mostrar.append(f"{nombre_imp}: {cur_snap.get('ts','?')}")

            if vista == "Mensual" and mes_key_sel is not None:
                # Find previous month snapshot
                sorted_keys = sorted(snapshots.keys())
                idx = sorted_keys.index(cur_key) if cur_key in sorted_keys else -1
                prev_key = sorted_keys[idx - 1] if idx > 0 else None
                prev_usuarios = {}
                if prev_key:
                    prev_usuarios = {u["usuario"]: u for u in snapshots[prev_key].get("usuarios", [])}

                for uname, u in cur_usuarios.items():
                    if uname in prev_usuarios:
                        pu = prev_usuarios[uname]
                        imp_bw    = max(0, u["imp_bw"]    - pu["imp_bw"])
                        imp_color = max(0, u["imp_color"] - pu["imp_color"])
                        cop_bw    = max(0, u["cop_bw"]    - pu["cop_bw"])
                        cop_color = max(0, u["cop_color"] - pu["cop_color"])
                    else:
                        imp_bw    = u["imp_bw"]
                        imp_color = u["imp_color"]
                        cop_bw    = u["cop_bw"]
                        cop_color = u["cop_color"]
                    total = imp_bw + imp_color + cop_bw + cop_color
                    nota = "" if prev_key else " (sin mes anterior)"
                    uid = u.get("id", "")
                    label = f"{uname} ({uid}){nota}" if uid else uname + nota
                    all_rows.append((nombre_imp, label, imp_bw, imp_color,
                                     cop_bw, cop_color, total))
            else:
                for u in cur_snap.get("usuarios", []):
                    uid = u.get("id", "")
                    label = f"{u['usuario']} ({uid})" if uid else u["usuario"]
                    all_rows.append((nombre_imp, label, u["imp_bw"], u["imp_color"],
                                     u["cop_bw"], u["cop_color"], u["total"]))

        # Separar en departamentos y usuarios
        def _nombre_base(label):
            # Quitar "(id)" y notas para comparar contra listas
            return label.split(" (")[0].lower().rstrip()

        deptos = [r for r in all_rows if _nombre_base(r[1]) in self._DEPARTAMENTOS]
        users  = [r for r in all_rows if _nombre_base(r[1]) not in self._DEPARTAMENTOS]

        _COL_IDX = {"impresora":0,"usuario":1,"imp_bw":2,"imp_color":3,
                    "cop_bw":4,"cop_color":5,"total":6}
        sort_idx = _COL_IDX.get(self._sort_col, 6)
        rev = self._sort_rev if self._sort_col else True

        def _sort_key(r):
            v = r[sort_idx]
            if isinstance(v, int): return v
            try: return int(v)
            except: return str(v).lower()

        deptos.sort(key=_sort_key, reverse=rev)
        users.sort(key=_sort_key, reverse=rev)

        # Totales globales
        tot_imp_bw    = sum(r[2] for r in all_rows)
        tot_imp_color = sum(r[3] for r in all_rows)
        tot_cop_bw    = sum(r[4] for r in all_rows)
        tot_cop_color = sum(r[5] for r in all_rows)
        tot_total     = sum(r[6] for r in all_rows)

        has_color = tot_imp_color > 0 or tot_cop_color > 0
        show_imp_col = (ip_filter == "Todas")
        self._update_imp_col_visibility(has_color)

        def _insertar_seccion(titulo, filas):
            if not filas:
                return
            collapsed = self._collapsed.get(titulo, False)
            icono = "▶" if collapsed else "▼"
            iid_sec = f"sec_{titulo}"
            if self.tree.exists(iid_sec):
                self.tree.delete(iid_sec)
            self.tree.insert("", "end", iid=iid_sec, values=(
                "", f"{icono}  {titulo}  ({len(filas)})",
                "", "", "", "", "",
            ), tags=("seccion",))
            if collapsed:
                return
            alt = False
            for r in filas:
                tag = "par" if alt else "impar"
                alt = not alt
                self.tree.insert("", "end", values=(
                    r[0] if show_imp_col else "",
                    r[1],
                    r[2] if r[2] else "—",
                    r[3] if r[3] else "—",
                    r[4] if r[4] else "—",
                    r[5] if r[5] else "—",
                    r[6] if r[6] else "—",
                ), tags=(tag,))

        _insertar_seccion("DEPARTAMENTOS", deptos)
        _insertar_seccion("USUARIOS", users)

        # Totals row
        if all_rows:
            self.tree.insert("", "end", values=(
                "" if not show_imp_col else "TOTAL",
                "TOTAL" if not show_imp_col else "",
                tot_imp_bw or "—", tot_imp_color or "—",
                tot_cop_bw or "—", tot_cop_color or "—",
                tot_total or "—",
            ), tags=("totales",))

        ts_str = "  |  ".join(ts_mostrar) if ts_mostrar else "Sin snapshots"
        self.lbl_info.configure(
            text=f"{len(all_rows)} usuarios  |  {ts_str}",
            text_color=TEXT2)

    def _exportar_excel(self):
        from tkinter.filedialog import asksaveasfilename

        if not self._datos:
            messagebox.showinfo("Sin datos", "No hay datos de contabilidad cargados.")
            return

        # Elegir impresoras a exportar
        ips_disp = list(self._datos.keys())
        nombres  = [self._datos[ip].get("nombre_impresora", ip) for ip in ips_disp]
        dlg = _DialogSeleccionarImpresoras(self, list(zip(ips_disp, nombres)))
        self.wait_window(dlg)
        if not dlg.resultado:
            return
        ips_sel = dlg.resultado

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            usa_xlsx = True
        except ImportError:
            usa_xlsx = False

        mes_label = self._sel_mes
        vista     = self._sel_vista
        mes_safe  = mes_label.replace(" ", "_")

        ext  = "xlsx" if usa_xlsx else "csv"
        ruta = asksaveasfilename(
            title="Guardar exportación",
            defaultextension=f".{ext}",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")] if usa_xlsx else [("CSV", "*.csv")],
            initialfile=f"contabilidad_{mes_safe}.{ext}",
        )
        if not ruta:
            return

        CABECERAS = ["Usuario", "ID", "Imp. B/N", "Imp. Color", "Cop. B/N", "Cop. Color", "Total"]

        def _filas_impresora(ip):
            bloque      = self._datos.get(ip, {})
            nombre_imp  = bloque.get("nombre_impresora", ip)
            snapshots   = bloque.get("snapshots", {})
            sorted_keys = sorted(snapshots.keys())

            if mes_label == "Acumulado" or vista == "Acumulado":
                cur_key  = sorted_keys[-1] if sorted_keys else None
                prev_key = None
            else:
                mk = self._mes_key_from_label(mes_label)
                cur_key  = mk if mk in snapshots else (sorted_keys[-1] if sorted_keys else None)
                idx      = sorted_keys.index(cur_key) if cur_key in sorted_keys else -1
                prev_key = sorted_keys[idx - 1] if idx > 0 else None

            if cur_key is None:
                return nombre_imp, [], []

            cur_u  = {u["usuario"]: u for u in snapshots[cur_key].get("usuarios", [])}
            prev_u = {u["usuario"]: u for u in snapshots[prev_key].get("usuarios", [])} if prev_key else {}

            deptos, users = [], []
            for uname, u in cur_u.items():
                uid = u.get("id", "")
                if prev_key and uname in prev_u:
                    pu = prev_u[uname]
                    ib = max(0, u["imp_bw"]    - pu["imp_bw"])
                    ic = max(0, u["imp_color"] - pu["imp_color"])
                    cb = max(0, u["cop_bw"]    - pu["cop_bw"])
                    cc = max(0, u["cop_color"] - pu["cop_color"])
                else:
                    ib, ic, cb, cc = u["imp_bw"], u["imp_color"], u["cop_bw"], u["cop_color"]
                fila = [uname, uid, ib, ic, cb, cc, ib+ic+cb+cc]
                if uname.lower().rstrip() in self._DEPARTAMENTOS:
                    deptos.append(fila)
                else:
                    users.append(fila)
            deptos.sort(key=lambda r: r[6], reverse=True)
            users.sort(key=lambda r: r[6], reverse=True)
            return nombre_imp, deptos, users

        try:
            if usa_xlsx:
                wb = openpyxl.Workbook()
                wb.remove(wb.active)  # quitar hoja vacía por defecto

                for ip in ips_sel:
                    nombre_imp, deptos, users = _filas_impresora(ip)
                    # Título de hoja: máx 31 chars, sin caracteres inválidos
                    sheet_name = "".join(c if c not in r'\/?*[]:'  else "_" for c in nombre_imp)[:31]
                    ws = wb.create_sheet(title=sheet_name)

                    # Fila de título
                    ws.append([f"{nombre_imp}  —  {mes_label}  ({vista})"])
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CABECERAS))
                    ws.cell(1, 1).font      = Font(bold=True, size=12)
                    ws.cell(1, 1).alignment = Alignment(horizontal="center")
                    ws.append([])

                    todas = deptos + users
                    for seccion, filas in [("DEPARTAMENTOS", deptos), ("USUARIOS", users)]:
                        if not filas:
                            continue
                        ws.append([f"{seccion} ({len(filas)})"])
                        ws.cell(ws.max_row, 1).font = Font(bold=True)
                        ws.append(CABECERAS)
                        for col in range(1, len(CABECERAS)+1):
                            ws.cell(ws.max_row, col).font = Font(bold=True)
                            ws.cell(ws.max_row, col).alignment = Alignment(horizontal="center")
                        for fila in filas:
                            ws.append(fila)
                            for col in range(3, len(CABECERAS)+1):
                                ws.cell(ws.max_row, col).alignment = Alignment(horizontal="center")
                        ws.append([])

                    # Fila TOTAL
                    ws.append(["TOTAL", "",
                               sum(r[2] for r in todas), sum(r[3] for r in todas),
                               sum(r[4] for r in todas), sum(r[5] for r in todas),
                               sum(r[6] for r in todas)])
                    for col in range(1, len(CABECERAS)+1):
                        ws.cell(ws.max_row, col).font = Font(bold=True)
                        if col > 2:
                            ws.cell(ws.max_row, col).alignment = Alignment(horizontal="center")

                    ws.column_dimensions["A"].width = 30
                    ws.column_dimensions["B"].width = 12
                    for letra in ("C","D","E","F","G"):
                        ws.column_dimensions[letra].width = 13

                wb.save(ruta)
            else:
                # Fallback: un CSV con separador de secciones por impresora
                import csv as _csv
                with open(ruta, "w", encoding="utf-8-sig", newline="") as f:
                    w = _csv.writer(f)
                    for ip in ips_sel:
                        nombre_imp, deptos, users = _filas_impresora(ip)
                        w.writerow([f"=== {nombre_imp} — {mes_label} ({vista}) ==="])
                        w.writerow([])
                        todas = deptos + users
                        for seccion, filas in [("DEPARTAMENTOS", deptos), ("USUARIOS", users)]:
                            if not filas:
                                continue
                            w.writerow([f"{seccion} ({len(filas)})"])
                            w.writerow(CABECERAS)
                            w.writerows(filas)
                            w.writerow([])
                        w.writerow(["TOTAL","",
                                    sum(r[2] for r in todas), sum(r[3] for r in todas),
                                    sum(r[4] for r in todas), sum(r[5] for r in todas),
                                    sum(r[6] for r in todas)])
                        w.writerow([]); w.writerow([])

            messagebox.showinfo("Exportación completada",
                f"✔  Archivo guardado:\n{ruta}")
        except Exception as e:
            messagebox.showerror("Error al exportar", str(e))

    def _limpiar(self):
        if messagebox.askyesno("Limpiar", "¿Eliminar todos los datos de contabilidad cargados?"):
            self._datos = {}
            guardar_json(CONTABILIDAD_FILE, self._datos)
            self._refresh_controls()
            self._poblar()

    def _on_tree_click(self, event):
        iid = self.tree.identify_row(event.y)
        if iid and iid.startswith("sec_"):
            titulo = iid[4:]  # quitar "sec_"
            self._collapsed[titulo] = not self._collapsed.get(titulo, False)
            self._poblar()

    def _sort(self, col):
        # Al ordenar, repoblar con el nuevo criterio para respetar secciones
        if self._sort_col == col:
            self._sort_rev = not self._sort_rev
        else:
            self._sort_col = col
            self._sort_rev = False
        self._poblar()


class _DialogCredencialesXSA(ctk.CTkToplevel):
    """Pide usuario/contraseña web y qué impresoras descargar."""
    def __init__(self, parent, impresoras):
        super().__init__(parent)
        self.resultado = None
        self.title("Descargar informe XSA")
        self.geometry("440x480")
        self.minsize(380, 400)
        self.configure(fg_color=BG2)
        self.grab_set(); self.lift(); self.focus_force()

        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=4, pady=4)

        pad = {"padx": 20, "pady": 5}

        ctk.CTkLabel(scroll, text="Usuario web de la impresora",
                     text_color=TEXT2, font=("Segoe UI", 11)).pack(anchor="w", **pad)
        self.e_user = ctk.CTkEntry(scroll, width=360, fg_color=BG3,
                                    border_color=BORDER, text_color=TEXT)
        self.e_user.insert(0, "admin")
        self.e_user.pack(**pad)

        ctk.CTkLabel(scroll, text="Contraseña",
                     text_color=TEXT2, font=("Segoe UI", 11)).pack(anchor="w", **pad)
        self.e_pass = ctk.CTkEntry(scroll, width=360, fg_color=BG3,
                                    border_color=BORDER, text_color=TEXT, show="•")
        self.e_pass.pack(**pad)

        ctk.CTkFrame(scroll, fg_color=BORDER, height=1).pack(fill="x", padx=20, pady=8)
        ctk.CTkLabel(scroll, text="Impresoras a descargar:",
                     text_color=TEXT2, font=("Segoe UI", 11)).pack(anchor="w", **pad)

        self._checks = {}
        for imp in impresoras:
            var = tk.BooleanVar(value=True)
            self._checks[imp["ip"]] = (var, imp)
            ctk.CTkCheckBox(scroll, text=f"{imp['nombre']}  ({imp['ip']})",
                            variable=var, text_color=TEXT,
                            font=("Segoe UI", 11),
                            fg_color=ACCENT).pack(anchor="w", padx=24, pady=3)

        btns = ctk.CTkFrame(self, fg_color=BG2, height=52)
        btns.pack(side="bottom", fill="x", padx=20, pady=10)
        btns.pack_propagate(False)
        ctk.CTkButton(btns, text="Cancelar", width=120, height=36,
                      fg_color=BG3, text_color=TEXT,
                      command=self.destroy).pack(side="left", padx=4)
        ctk.CTkButton(btns, text="⟳  Descargar", width=140, height=36,
                      fg_color=ACCENT, hover_color="#3a7de8", text_color=TEXT,
                      font=("Segoe UI", 12, "bold"),
                      command=self._ok).pack(side="right", padx=4)

    def _ok(self):
        pwd = self.e_pass.get().strip()
        usr = self.e_user.get().strip() or "admin"
        if not pwd:
            return
        self.resultado = [
            {"ip": ip, "nombre": imp["nombre"], "usuario": usr, "password": pwd}
            for ip, (var, imp) in self._checks.items()
            if var.get()
        ]
        self.destroy()


class _DialogSeleccionarImpresoras(ctk.CTkToplevel):
    """Checkboxes para elegir qué impresoras exportar."""
    def __init__(self, parent, opciones):  # opciones: [(ip, nombre), ...]
        super().__init__(parent)
        self.resultado = None
        self.title("Exportar — selecciona impresoras")
        self.geometry("420x300")
        self.configure(fg_color=BG2)
        self.grab_set(); self.lift(); self.focus_force()

        ctk.CTkLabel(self, text="Selecciona las impresoras a exportar:",
                     font=("Segoe UI", 12), text_color=TEXT).pack(pady=(16, 8))

        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=16)

        self._checks = {}
        for ip, nombre in opciones:
            var = tk.BooleanVar(value=True)
            self._checks[ip] = var
            ctk.CTkCheckBox(scroll, text=nombre, variable=var,
                            text_color=TEXT, font=("Segoe UI", 11),
                            fg_color=ACCENT).pack(anchor="w", pady=3)

        btns = ctk.CTkFrame(self, fg_color=BG2, height=52)
        btns.pack(side="bottom", fill="x", padx=16, pady=10)
        btns.pack_propagate(False)
        ctk.CTkButton(btns, text="Cancelar", width=100, height=34,
                      fg_color=BG3, text_color=TEXT,
                      command=self.destroy).pack(side="left")
        ctk.CTkButton(btns, text="↑  Exportar", width=120, height=34,
                      fg_color=ACCENT, hover_color="#3a7de8", text_color=TEXT,
                      font=("Segoe UI", 12, "bold"),
                      command=self._ok).pack(side="right")

    def _ok(self):
        self.resultado = [ip for ip, var in self._checks.items() if var.get()]
        if not self.resultado:
            return
        self.destroy()


class _DialogSeleccionarImpresora(ctk.CTkToplevel):
    def __init__(self, parent, opciones):
        super().__init__(parent)
        self.resultado = None
        self.title("¿A qué impresora pertenece este CSV?")
        self.geometry("400x280")
        self.configure(fg_color=BG2)
        self.grab_set(); self.lift(); self.focus_force()

        ctk.CTkLabel(self, text="Selecciona la impresora origen del CSV:",
                     font=("Segoe UI", 12), text_color=TEXT).pack(pady=(20, 8))

        self.var = tk.StringVar(value=opciones[0] if opciones else "")
        if opciones:
            for op in opciones:
                ctk.CTkRadioButton(self, text=op, variable=self.var, value=op,
                                   text_color=TEXT, font=("Segoe UI", 11)).pack(anchor="w", padx=30, pady=3)
        else:
            ctk.CTkEntry(self, textvariable=self.var,
                         placeholder_text="IP de la impresora",
                         width=300, fg_color=BG3, border_color=BORDER,
                         text_color=TEXT).pack(pady=8)

        ctk.CTkButton(self, text="✔  Confirmar", width=140, height=34,
                      fg_color=ACCENT, text_color=TEXT,
                      font=("Segoe UI", 12, "bold"),
                      command=self._ok).pack(pady=(16, 4))

    def _ok(self):
        self.resultado = self.var.get().strip()
        self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Fleet Monitor Pro")
        self.geometry("1280x720")
        self.minsize(960, 560)
        # Maximizar: zoomed (Windows/Linux), wm_attributes fallback
        try:
            self.after(100, lambda: self.wm_state("zoomed"))
        except Exception:
            try:
                self.after(100, lambda: self.wm_attributes("-zoomed", True))
            except Exception:
                pass
        self.configure(fg_color=BG)
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.cfg        = {**DEFAULT_CONFIG, **cargar_json(CONFIG_FILE, {})}
        self.impresoras = cargar_impresoras()
        self.cache      = self._cargar_cache_historial()
        self.sel_ip     = None
        self._scan_thread = None
        self._autoref_job = None
        self._reminder_shown = False

        # Precargar contraseña por defecto si no está guardada
        if not self.cfg.get("xsa_password"):
            self.cfg["xsa_password"] = "1111"
            guardar_json(CONFIG_FILE, self.cfg)

        self._web_proc = self._arrancar_web_server()

        self._build_ui()
        self._poblar_tabla()
        self._schedule_autoref()
        self._schedule_xsa_autodownload()
        self._check_monthly_reminder()
        self._setup_tray()

        if not SNMP_OK:
            messagebox.showerror("Error SNMP", f"No se pudo cargar pysnmp:\n{SNMP_ERROR}\n\nEjecuta: pip install pysnmp")

    # ── WEB SERVER ────────────────────────────────────────────────────────────
    def _arrancar_web_server(self):
        """Arranca web_server.py como subproceso oculto sin ventana de consola."""
        script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "web_server.py")
        if not os.path.exists(script):
            return None
        kwargs = {}
        if sys.platform == "win32":
            kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
        try:
            return subprocess.Popen([sys.executable, script], **kwargs)
        except Exception:
            return None

    # ── SYSTEM TRAY ───────────────────────────────────────────────────────────
    def _setup_tray(self):
        if not TRAY_OK or not PIL_OK:
            # Sin pystray: cerrar ventana termina la app normalmente
            return

        # Crear icono 64×64 con la letra "F" (Fleet)
        img = PILImage.new("RGBA", (64, 64), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        draw.ellipse([0, 0, 63, 63], fill="#4f8ef7")
        draw.text((18, 14), "F", fill="white")

        menu = pystray.Menu(
            pystray.MenuItem("Mostrar Fleet Monitor", self._tray_mostrar, default=True),
            pystray.Menu.SEPARATOR,
            pystray.MenuItem("Salir", self._tray_salir),
        )
        self._tray_icon = pystray.Icon("FleetMonitor", img, "Fleet Monitor Pro", menu)

        # Interceptar el botón X → minimizar a bandeja
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        # Arrancar el icono en su propio hilo
        t = threading.Thread(target=self._tray_icon.run, daemon=True)
        t.start()

    def _on_close(self):
        """Cerrar ventana → ocultar a bandeja."""
        self.withdraw()

    def _tray_mostrar(self, icon=None, item=None):
        """Doble clic o "Mostrar" → restaurar ventana."""
        self.after(0, self._restaurar_ventana)

    def _restaurar_ventana(self):
        self.deiconify()
        self.lift()
        self.focus_force()
        try:
            self.wm_state("zoomed")
        except Exception:
            pass

    def _tray_salir(self, icon=None, item=None):
        """Salir de verdad desde el menú de bandeja."""
        if TRAY_OK and hasattr(self, "_tray_icon"):
            self._tray_icon.stop()
        if hasattr(self, "_web_proc") and self._web_proc and self._web_proc.poll() is None:
            self._web_proc.terminate()
        self.after(0, self.destroy)

    # ── BUILD UI ──────────────────────────────────────────────────────────────
    def _cargar_cache_historial(self):
        """Precarga la caché con el último registro guardado de cada impresora."""
        historial = cargar_json(HISTORIAL_FILE, {})
        cache = {}
        for imp in self.impresoras:
            ip = imp["ip"]
            registros = historial.get(ip, [])
            if registros:
                ultimo = registros[-1]
                ts_str = ultimo.get("ts", "")
                try:
                    ts = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
                except Exception:
                    ts = None
                cache[ip] = {
                    "consumibles": ultimo.get("consumibles"),
                    "error":       None,
                    "ts":          ts,
                    "estado":      "ok",   # se actualizará en el primer scan
                    "info":        {},
                }
        return cache

    def _build_ui(self):
        # ── Menubar nativo ──
        menubar = tk.Menu(self, bg=BG2, fg=TEXT, activebackground=ACCENT,
                          activeforeground=TEXT, tearoff=False)

        m_disp = tk.Menu(menubar, tearoff=False, bg=BG2, fg=TEXT,
                         activebackground=ACCENT, activeforeground=TEXT)
        m_disp.add_command(label="Añadir manualmente", command=self._añadir)
        m_disp.add_command(label="Escanear red...", command=self._escanear_red)
        m_disp.add_separator()
        m_disp.add_command(label="Editar", command=self._editar)
        m_disp.add_command(label="Eliminar", command=self._eliminar)
        menubar.add_cascade(label="Dispositivos", menu=m_disp)

        m_tools = tk.Menu(menubar, tearoff=False, bg=BG2, fg=TEXT,
                          activebackground=ACCENT, activeforeground=TEXT)
        m_tools.add_command(label="Exportar...", command=self._exportar_todo)
        m_tools.add_separator()
        m_tools.add_command(label="Contabilidad por usuario...", command=self._abrir_contabilidad)
        m_tools.add_separator()
        m_tools.add_command(label="Configuración", command=self._abrir_config)
        menubar.add_cascade(label="Herramientas", menu=m_tools)

        m_help = tk.Menu(menubar, tearoff=False, bg=BG2, fg=TEXT,
                         activebackground=ACCENT, activeforeground=TEXT)
        m_help.add_command(label="Acerca de...", command=self._acerca_de)
        menubar.add_cascade(label="Ayuda", menu=m_help)

        self.configure(menu=menubar)

        # ── Toolbar slim ──
        tb = ctk.CTkFrame(self, fg_color=BG2, height=48, corner_radius=0)
        tb.pack(fill="x")
        tb.pack_propagate(False)

        self.btn_volver = ctk.CTkButton(tb, text="←  Volver", width=100, height=32,
                      fg_color=BG3, hover_color=BORDER, text_color=TEXT,
                      font=("Segoe UI", 11),
                      command=self._volver_inicio)
        # se muestra solo en vistas secundarias

        ctk.CTkLabel(tb, text="🖨 Fleet Monitor Pro",
                     font=("Segoe UI", 14, "bold"), text_color=TEXT).pack(side="left", padx=16)
        self.lbl_ts = ctk.CTkLabel(tb, text="", font=("Segoe UI", 10), text_color=TEXT2)
        self.lbl_ts.pack(side="left", padx=6)

        self.btn_refrescar = ctk.CTkButton(tb, text="↺  Refrescar", width=110, height=32,
                      fg_color=ACCENT, hover_color="#3a7de8",
                      text_color=TEXT, font=("Segoe UI", 11, "bold"),
                      command=self._refrescar_todo)
        self.btn_refrescar.pack(side="right", padx=12, pady=8)

        # Contenedor de vistas (inicio / contabilidad / ...)
        self._views = ctk.CTkFrame(self, fg_color=BG, corner_radius=0)
        self._views.pack(fill="both", expand=True)

        self._view_inicio = ctk.CTkFrame(self._views, fg_color=BG, corner_radius=0)
        self._view_contabilidad = ctk.CTkFrame(self._views, fg_color=BG, corner_radius=0)

        self._build_ui_rest()
        self._build_contabilidad_embebida()
        self._mostrar_vista(self._view_inicio)

    def _mostrar_vista(self, vista):
        for v in (self._view_inicio, self._view_contabilidad):
            v.pack_forget()
        vista.pack(fill="both", expand=True)
        es_inicio = (vista is self._view_inicio)
        if es_inicio:
            self.btn_volver.pack_forget()
            self.btn_refrescar.pack(side="right", padx=12, pady=8)
        else:
            self.btn_refrescar.pack_forget()
            self.btn_volver.pack(side="left", padx=6, pady=8)

    def _volver_inicio(self):
        self._mostrar_vista(self._view_inicio)

    def _build_contabilidad_embebida(self):
        self._dlg_contabilidad = DialogContabilidadEmbebida(self._view_contabilidad, self)
        self._dlg_contabilidad.pack(fill="both", expand=True)

    def _abrir_contabilidad(self):
        # Refrescar datos por si han cambiado
        self._dlg_contabilidad._datos = cargar_json(CONTABILIDAD_FILE, {})
        self._dlg_contabilidad._migrar_formato_antiguo()
        self._dlg_contabilidad._refresh_controls()
        self._dlg_contabilidad._poblar()
        self._mostrar_vista(self._view_contabilidad)

    def _acerca_de(self):
        messagebox.showinfo("Acerca de Fleet Monitor Pro",
                            "Fleet Monitor Pro\nVersión 1.0\n\nMonitorización de consumibles de impresoras vía SNMP.")

    def _check_monthly_reminder(self):
        if self._reminder_shown:
            return
        now = datetime.now()
        if now.day != 1:
            return
        mes_actual = now.strftime("%Y-%m")
        datos = cargar_json(CONTABILIDAD_FILE, {})
        if not datos:
            return
        for ip_key, bloque in datos.items():
            snapshots = bloque.get("snapshots", {})
            if mes_actual not in snapshots:
                self._reminder_shown = True
                messagebox.showinfo(
                    "Recordatorio mensual",
                    "Recordatorio: Hoy es el 1 de mes. Descarga el informe XSA para registrar "
                    "el consumo mensual (Herramientas → Contabilidad).")
                return

    def _build_ui_rest(self):
        vi = self._view_inicio
        # ── KPI bar ──
        kpi_bar = ctk.CTkFrame(vi, fg_color=BG2, height=56, corner_radius=0)
        kpi_bar.pack(fill="x")
        kpi_bar.pack_propagate(False)
        sep = ctk.CTkFrame(vi, fg_color=BORDER, height=1, corner_radius=0)
        sep.pack(fill="x")

        self.kpi_vars = {}
        for key, lbl, color in [
            ("online",   "En línea",      OK),
            ("offline",  "Sin conexión",  OFFLINE),
            ("criticos", "Nivel crítico", CRIT),
            ("alertas",  "En alerta",     WARN),
        ]:
            f = ctk.CTkFrame(kpi_bar, fg_color="transparent")
            f.pack(side="left", padx=28, pady=6)
            var = tk.StringVar(value="0")
            self.kpi_vars[key] = var
            ctk.CTkLabel(f, textvariable=var, font=("Segoe UI", 20, "bold"), text_color=color).pack()
            ctk.CTkLabel(f, text=lbl, font=("Segoe UI", 9), text_color=TEXT2).pack()

        # ── Search bar ──
        sb = ctk.CTkFrame(vi, fg_color=BG, height=38, corner_radius=0)
        sb.pack(fill="x", padx=8, pady=(6,2))
        sb.pack_propagate(False)

        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self._filtrar())
        ctk.CTkEntry(sb, textvariable=self.search_var,
                     placeholder_text="🔍  Buscar por nombre o IP...",
                     width=300, height=30, fg_color=BG2, border_color=BORDER, text_color=TEXT,
                     font=("Segoe UI", 11)).pack(side="left", padx=4)

        self.solo_alertas_var = tk.BooleanVar()
        ctk.CTkCheckBox(sb, text="Solo con alertas", variable=self.solo_alertas_var,
                        text_color=TEXT2, font=("Segoe UI", 11),
                        command=self._filtrar).pack(side="left", padx=14)

        self.lbl_scan = ctk.CTkLabel(sb, text="", font=("Segoe UI", 10), text_color=TEXT2)
        self.lbl_scan.pack(side="left", padx=10)

        # ── Cuerpo: tabla + panel ──
        body = ctk.CTkFrame(vi, fg_color=BG, corner_radius=0)
        body.pack(fill="both", expand=True, padx=8, pady=(0,8))

        # Tabla
        tbl_frame = ctk.CTkFrame(body, fg_color=BG2, corner_radius=6)
        tbl_frame.pack(side="left", fill="both", expand=True, padx=(0,4))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("F.Treeview",
            background=BG2, fieldbackground=BG2, foreground=TEXT,
            rowheight=30, borderwidth=0, font=("Segoe UI", 10))
        style.configure("F.Treeview.Heading",
            background=BG3, foreground=TEXT2,
            font=("Segoe UI", 9, "bold"), relief="flat", borderwidth=0)
        style.map("F.Treeview",
            background=[("selected", ROW_SEL)],
            foreground=[("selected", TEXT)])
        style.layout("F.Treeview", [("F.Treeview.treearea", {"sticky": "nswe"})])

        cols = ("estado","nombre","ip","ubicacion","paginas","consumibles","ts")
        self.tree = ttk.Treeview(tbl_frame, columns=cols, show="headings",
                                  style="F.Treeview", selectmode="browse")
        for col, hdr, w, stretch in [
            ("estado",     "Estado",       80,  False),
            ("nombre",     "Nombre",       190, True),
            ("ip",         "IP",           115, False),
            ("ubicacion",  "Ubicación",    120, False),
            ("paginas",    "Páginas",      75,  False),
            ("consumibles","Consumibles",  280, True),
            ("ts",         "Última lect.", 75,  False),
        ]:
            self.tree.heading(col, text=hdr,
                command=lambda c=col: self._sort(c))
            self.tree.column(col, width=w, anchor="w", stretch=stretch)

        self.tree.tag_configure("ok",      background=BG2,      foreground=TEXT)
        self.tree.tag_configure("alerta",  background="#29200a", foreground=WARN)
        self.tree.tag_configure("critico", background="#280a0a", foreground=CRIT)
        self.tree.tag_configure("offline", background=BG2,      foreground=OFFLINE)
        self.tree.tag_configure("alt_ok",  background=ROW_ALT,  foreground=TEXT)

        vsb = ttk.Scrollbar(tbl_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Double-1>", lambda e: self._refrescar_sel())

        # Panel detalle
        self.panel = ctk.CTkFrame(body, fg_color=BG2, corner_radius=6, width=290)
        self.panel.pack(side="right", fill="y", padx=(4,0))
        self.panel.pack_propagate(False)
        self._panel_vacio()

        self._sort_col = None
        self._sort_rev = False

    # ── PANEL DETALLE ─────────────────────────────────────────────────────────
    def _panel_vacio(self):
        for w in self.panel.winfo_children(): w.destroy()
        ctk.CTkLabel(self.panel, text="🖨️", font=("Segoe UI", 36), text_color=TEXT2).pack(pady=(70,8))
        ctk.CTkLabel(self.panel, text="Selecciona un dispositivo\npara ver el detalle",
                     font=("Segoe UI", 12), text_color=TEXT2, justify="center").pack()

    def _panel_detalle(self, ip):
        for w in self.panel.winfo_children(): w.destroy()
        imp   = next((i for i in self.impresoras if i["ip"]==ip), None)
        if not imp: return
        cache = self.cache.get(ip, {})
        cons  = cache.get("consumibles")
        error = cache.get("error")
        ts    = cache.get("ts")
        estado= cache.get("estado","offline")

        color_map  = {"ok":OK,"alerta":WARN,"critico":CRIT,"offline":OFFLINE}
        label_map  = {"ok":"OK","alerta":"Alerta","critico":"Crítico","offline":"Offline"}
        color = color_map.get(estado, OFFLINE)
        label = label_map.get(estado, "—")

        # Todo el contenido dentro de un único scroll para que nada quede cortado
        sc = ctk.CTkScrollableFrame(self.panel, fg_color="transparent")
        sc.pack(fill="both", expand=True, padx=0, pady=0)

        # ── A) Imagen del dispositivo ──
        img_frame = ctk.CTkFrame(sc, fg_color=BG3, corner_radius=6,
                                  width=270, height=110)
        img_frame.pack(fill="x", padx=10, pady=(10, 4))
        img_frame.pack_propagate(False)

        imagen_path = imp.get("imagen", "")
        imagen_cargada = False
        if PIL_OK and imagen_path and os.path.isfile(imagen_path):
            try:
                pil_img = PILImage.open(imagen_path)
                pil_img.thumbnail((260, 100), PILImage.LANCZOS)
                ctk_img = ctk.CTkImage(light_image=pil_img, dark_image=pil_img,
                                        size=(pil_img.width, pil_img.height))
                lbl_img = ctk.CTkLabel(img_frame, image=ctk_img, text="")
                lbl_img.pack(expand=True)
                imagen_cargada = True
            except Exception:
                pass

        if not imagen_cargada:
            ctk.CTkLabel(img_frame, text="🖨", font=("Segoe UI", 36),
                         text_color=TEXT2).pack(expand=True, pady=(10, 0))
            info_cache = cache.get("info", {})
            modelo_short = (info_cache.get("modelo") or imp.get("nombre", ""))[:30]
            ctk.CTkLabel(img_frame, text=modelo_short, font=("Segoe UI", 9),
                         text_color=TEXT2).pack(pady=(0, 8))

        # ── B) Header: nombre + badge estado + IP ──
        hdr = ctk.CTkFrame(sc, fg_color=BG3, corner_radius=6)
        hdr.pack(fill="x", padx=10, pady=(4, 4))
        top = ctk.CTkFrame(hdr, fg_color="transparent")
        top.pack(fill="x", padx=10, pady=(8, 2))
        ctk.CTkLabel(top, text=imp["nombre"], font=("Segoe UI", 12, "bold"),
                     text_color=TEXT, wraplength=170, justify="left").pack(side="left", anchor="w")
        ctk.CTkLabel(top, text=f"● {label}", font=("Segoe UI", 10, "bold"),
                     text_color=color).pack(side="right")
        ctk.CTkLabel(hdr, text=ip, font=("Consolas", 10), text_color=TEXT2).pack(anchor="w", padx=10, pady=(0, 8))

        # ── C) Botones de acción rápida ──
        acc = ctk.CTkFrame(sc, fg_color="transparent")
        acc.pack(fill="x", padx=10, pady=(2, 4))
        ctk.CTkButton(acc, text="🌐 Web", width=80, height=28,
                      fg_color=ACCENT, hover_color="#3a7de8", text_color=TEXT,
                      font=("Segoe UI", 10),
                      command=lambda: webbrowser.open(f"http://{ip}")).pack(side="left", padx=(0, 4))
        ctk.CTkButton(acc, text="↺ Refrescar", width=90, height=28,
                      fg_color=BG3, hover_color=BORDER, text_color=TEXT,
                      font=("Segoe UI", 10),
                      command=self._refrescar_sel).pack(side="left", padx=(0, 4))
        ctk.CTkButton(acc, text="↓ CSV", width=70, height=28,
                      fg_color=BG3, hover_color=BORDER, text_color=TEXT,
                      font=("Segoe UI", 10),
                      command=lambda: self._exportar_csv(ip)).pack(side="left")

        # Helpers anclados al scroll
        def fila(lbl, val):
            f = ctk.CTkFrame(sc, fg_color="transparent")
            f.pack(fill="x", padx=14, pady=1)
            ctk.CTkLabel(f, text=lbl, font=("Segoe UI", 10), text_color=TEXT2, width=75, anchor="w").pack(side="left")
            ctk.CTkLabel(f, text=val, font=("Segoe UI", 10, "bold"), text_color=TEXT, anchor="w", wraplength=165).pack(side="left")

        def sep_line():
            ctk.CTkFrame(sc, fg_color=BORDER, height=1, corner_radius=0).pack(fill="x", padx=10, pady=(6, 2))

        def seccion(titulo):
            ctk.CTkLabel(sc, text=titulo, font=("Segoe UI", 8, "bold"),
                         text_color=TEXT2).pack(anchor="w", padx=14, pady=(2, 2))

        # ── D) Sección INFORMACIÓN ──
        sep_line()
        seccion("INFORMACIÓN")
        fila("Ubicación:", imp.get("ubicacion", "") or "—")
        fila("Comunidad:", imp.get("comunidad", "") or self.cfg["comunidad_snmp"])
        fila("Lectura:",   ts.strftime("%H:%M:%S") if ts else "—")

        # ── E) Sección DISPOSITIVO ──
        info = cache.get("info", {})
        if info:
            sep_line()
            seccion("DISPOSITIVO")
            if info.get("modelo"):
                fila("Modelo:", info["modelo"][:35])
            if info.get("sys_nombre"):
                fila("Hostname:", info["sys_nombre"][:35])
            if info.get("serial"):
                fila("Serie:", info["serial"][:25])
            if info.get("paginas"):
                try:   pag_fmt = f"{int(info['paginas']):,}".replace(",", ".")
                except: pag_fmt = info["paginas"]
                fila("Páginas:", pag_fmt)
            if info.get("estado_imp"):
                fila("Estado imp.:", ESTADO_IMP_MAP.get(info["estado_imp"], info["estado_imp"]))
            if info.get("bandeja_nivel") and info.get("bandeja_cap"):
                try:
                    bn = int(info["bandeja_nivel"])
                    bc = int(info["bandeja_cap"])
                    if bc > 0:
                        fila("Bandeja:", f"{bn}/{bc} hojas ({round(bn/bc*100)}%)")
                except: pass
            if info.get("sys_uptime"):
                fila("Uptime:", _formatear_uptime(info["sys_uptime"]))

        # ── F) Sección ALERTAS ──
        alerts = cache.get("alerts") or []
        if alerts:
            sep_line()
            seccion("ALERTAS ACTIVAS")
            for a in alerts:
                color_a = CRIT if a["nivel"] == "critical" else WARN if a["nivel"] == "warning" else TEXT2
                f = ctk.CTkFrame(sc, fg_color="transparent")
                f.pack(fill="x", padx=14, pady=1)
                ctk.CTkLabel(f, text=a["icono"], font=("Segoe UI", 11),
                             text_color=color_a, width=20).pack(side="left")
                ctk.CTkLabel(f, text=a["texto"], font=("Segoe UI", 10),
                             text_color=color_a, anchor="w", wraplength=210).pack(side="left", padx=(4, 0))

        # ── G) Sección CONSUMIBLES con barras ──
        sep_line()
        seccion("CONSUMIBLES")

        if error:
            ctk.CTkLabel(sc, text=f"⚠  {error}", font=("Segoe UI", 10),
                         text_color=WARN, wraplength=240, justify="left").pack(anchor="w", padx=14, pady=6)
        elif cons:
            for c in cons:
                pct = c["porcentaje"]
                bc  = CRIT if pct<=self.cfg["umbral_critico"] else WARN if pct<=self.cfg["umbral_alerta"] else OK
                ctk.CTkLabel(sc, text=c["componente"], font=("Segoe UI", 10),
                             text_color=TEXT, anchor="w").pack(fill="x", padx=14, pady=(6, 0))
                bar_bg = ctk.CTkFrame(sc, fg_color=BG3, height=10, corner_radius=5)
                bar_bg.pack(fill="x", padx=14, pady=(2, 0))
                bar_bg.pack_propagate(False)
                if pct > 0:
                    ctk.CTkFrame(bar_bg, fg_color=bc, height=10, corner_radius=5,
                                 width=int(pct/100*230)).place(x=0, y=0, relheight=1)
                row_pct = ctk.CTkFrame(sc, fg_color="transparent")
                row_pct.pack(fill="x", padx=14)
                ctk.CTkLabel(row_pct, text=f"{pct}%", font=("Segoe UI", 9, "bold"),
                             text_color=bc).pack(side="right")
        else:
            ctk.CTkLabel(sc, text="Sin datos disponibles", text_color=TEXT2,
                         font=("Segoe UI", 10)).pack(anchor="w", padx=14, pady=6)
        # Espaciado final para que el último elemento no quede pegado al borde
        ctk.CTkFrame(sc, fg_color="transparent", height=10).pack()

    # ── TABLA ─────────────────────────────────────────────────────────────────
    def _poblar_tabla(self):
        filtro = self.search_var.get().lower()
        solo   = self.solo_alertas_var.get()
        for row in self.tree.get_children():
            self.tree.delete(row)

        kpi = {"online":0,"offline":0,"criticos":0,"alertas":0}
        alt = False

        for imp in self.impresoras:
            ip = imp["ip"]
            if filtro and filtro not in ip.lower() and filtro not in imp["nombre"].lower():
                continue

            cache  = self.cache.get(ip, {})
            cons   = cache.get("consumibles")
            error  = cache.get("error")
            ts     = cache.get("ts")
            estado = cache.get("estado","pendiente")
            ts_str = ts.strftime("%H:%M") if ts else "—"

            if estado == "offline":   kpi["offline"]  += 1
            elif estado == "ok":      kpi["online"]   += 1
            elif estado == "alerta":  kpi["online"]   += 1;  kpi["alertas"]  += 1
            elif estado == "critico": kpi["online"]   += 1;  kpi["criticos"] += 1

            if solo and estado in ("ok","pendiente"):
                continue

            if error:
                cons_str = f"⚠  {error[:55]}"
            elif cons:
                cons_str = _formato_consumibles(cons, self.cfg["umbral_critico"], self.cfg["umbral_alerta"])
            elif estado == "pendiente":
                cons_str = "Pendiente de escaneo..."
            else:
                cons_str = "—"

            dot  = {"ok":"●","alerta":"●","critico":"●","offline":"○","pendiente":"…"}.get(estado,"?")
            elbl = {"ok":"OK","alerta":"Alerta","critico":"Crítico","offline":"Offline","pendiente":"—"}.get(estado,"")
            alerts = cache.get("alerts") or []
            n_crit = sum(1 for a in alerts if a["nivel"] == "critical")
            n_warn = sum(1 for a in alerts if a["nivel"] == "warning")
            badge = ""
            if n_crit:  badge += f"  🔴{n_crit}"
            if n_warn:  badge += f"  🟡{n_warn}"
            estado_str = f"{dot}  {elbl}{badge}"

            tag = estado if estado in ("ok","alerta","critico","offline") else "ok"
            if alt and tag == "ok":
                tag = "alt_ok"
            alt = not alt

            paginas_str = "—"
            if cache.get("info"):
                pag = cache["info"].get("paginas","")
                if pag:
                    try: paginas_str = f"{int(pag):,}".replace(",",".")
                    except: paginas_str = pag

            # Usar hostname SNMP si está disponible, sino el nombre configurado
            info_cache = cache.get("info", {})
            nombre_mostrar = info_cache.get("sys_nombre") or imp["nombre"]

            self.tree.insert("", "end", iid=ip, values=(
                estado_str, nombre_mostrar, ip,
                imp.get("ubicacion",""), paginas_str, cons_str, ts_str
            ), tags=(tag,))

        for k, v in kpi.items():
            self.kpi_vars[k].set(str(v))
        self.lbl_ts.configure(text=f"  {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

    def _filtrar(self):
        self._poblar_tabla()

    def _sort(self, col):
        if self._sort_col == col:
            self._sort_rev = not self._sort_rev
        else:
            self._sort_col = col
            self._sort_rev = False
        idx = {"estado":0,"nombre":1,"ip":2,"ubicacion":3,"paginas":4,"consumibles":5,"ts":6}[col]
        rows = [(self.tree.set(k, col), k) for k in self.tree.get_children()]
        rows.sort(reverse=self._sort_rev)
        for i, (_, k) in enumerate(rows):
            self.tree.move(k, "", i)

    def _on_select(self, event):
        sel = self.tree.selection()
        if sel:
            self.sel_ip = sel[0]
            self._panel_detalle(self.sel_ip)

    # ── ESCANEO ───────────────────────────────────────────────────────────────
    def _refrescar_todo(self):
        self.cache.clear()
        self._poblar_tabla()
        if self._scan_thread and self._scan_thread.is_alive():
            return
        self._scan_thread = threading.Thread(target=self._scan_worker, daemon=True)
        self._scan_thread.start()

    def _refrescar_sel(self):
        if not self.sel_ip: return
        self.cache.pop(self.sel_ip, None)
        ip = self.sel_ip
        def worker():
            imp = next((i for i in self.impresoras if i["ip"]==ip), {})
            com = imp.get("comunidad") or self.cfg["comunidad_snmp"]
            cons, err = obtener_consumibles(ip, com, self.cfg["timeout_snmp"])
            info = obtener_info_extra(ip, com, self.cfg["timeout_snmp"]) if not err else {}
            alerts = obtener_alertas(ip, com, self.cfg["timeout_snmp"]) if not err else []
            if err:
                est = "offline"
            else:
                tc = any(c["porcentaje"]<=self.cfg["umbral_critico"] for c in (cons or []))
                ta = any(c["porcentaje"]<=self.cfg["umbral_alerta"]  for c in (cons or []))
                est = "critico" if tc else "alerta" if ta else "ok"
                if cons: guardar_historial(ip, cons, nombre=info.get("sys_nombre") or imp.get("nombre"))
            self.cache[ip] = {"consumibles":cons,"error":err,"ts":datetime.now(),"estado":est,"info":info,"alerts":alerts}
            self.after(0, self._poblar_tabla)
            if self.sel_ip == ip:
                self.after(0, lambda: self._panel_detalle(ip))
        threading.Thread(target=worker, daemon=True).start()

    def _scan_worker(self):
        total = len(self.impresoras)
        for i, imp in enumerate(self.impresoras):
            ip  = imp["ip"]
            com = imp.get("comunidad") or self.cfg["comunidad_snmp"]
            self.after(0, lambda t=f"Escaneando {i+1}/{total}: {ip}":
                       self.lbl_scan.configure(text=t))
            cons, err = obtener_consumibles(ip, com, self.cfg["timeout_snmp"])
            info = obtener_info_extra(ip, com, self.cfg["timeout_snmp"]) if not err else {}
            alerts = obtener_alertas(ip, com, self.cfg["timeout_snmp"]) if not err else []
            if err:
                est = "offline"
            else:
                tc = any(c["porcentaje"]<=self.cfg["umbral_critico"] for c in (cons or []))
                ta = any(c["porcentaje"]<=self.cfg["umbral_alerta"]  for c in (cons or []))
                est = "critico" if tc else "alerta" if ta else "ok"
                if cons: guardar_historial(ip, cons, nombre=info.get("sys_nombre") or imp.get("nombre"))
            self.cache[ip] = {"consumibles":cons,"error":err,"ts":datetime.now(),"estado":est,"info":info,"alerts":alerts}
            self.after(0, self._poblar_tabla)
            if self.sel_ip == ip:
                self.after(0, lambda _ip=ip: self._panel_detalle(_ip))
        self.after(0, lambda: self.lbl_scan.configure(text="✔  Escaneo completado"))

    # ── AUTO-REFRESCO ─────────────────────────────────────────────────────────
    def _schedule_autoref(self):
        if self._autoref_job:
            self.after_cancel(self._autoref_job)
            self._autoref_job = None
        seg = self.cfg.get("autorefresh_seg", 0)
        if seg > 0:
            self._autoref_job = self.after(seg * 1000, self._tick_autoref)

    def _tick_autoref(self):
        # No borrar caché: los datos anteriores permanecen hasta que lleguen los nuevos
        if self._scan_thread and self._scan_thread.is_alive():
            self._schedule_autoref()
            return
        self._scan_thread = threading.Thread(target=self._scan_worker, daemon=True)
        self._scan_thread.start()
        self._schedule_autoref()

    # ── XSA AUTO DOWNLOAD ─────────────────────────────────────────────────────
    def _schedule_xsa_autodownload(self):
        """Programa el próximo tick para cuando sea el día 1 a las 00:00."""
        if not self.cfg.get("xsa_autodownload") or not self.cfg.get("xsa_password"):
            return
        now = datetime.now()
        # Próximo día 1 a las 00:00
        if now.month == 12:
            prox = datetime(now.year + 1, 1, 1)
        else:
            prox = datetime(now.year, now.month + 1, 1)
        ms = int((prox - now).total_seconds() * 1000)
        self.after(ms, self._tick_xsa_autodownload)

    def _tick_xsa_autodownload(self):
        mes_actual = datetime.now().strftime("%Y-%m")
        if self.cfg.get("xsa_ultimo_mes") == mes_actual:
            self._schedule_xsa_autodownload()
            return
        if not REQUESTS_OK:
            self._schedule_xsa_autodownload()
            return

        usr = self.cfg.get("xsa_usuario", "admin")
        pwd = self.cfg.get("xsa_password", "")

        def worker():
            errores = []
            ok = 0
            datos = cargar_json(CONTABILIDAD_FILE, {})
            for imp in self.impresoras:
                ip = imp["ip"]
                csv_txt, err = _xsa_descargar_csv(ip, pwd, usr)
                if err:
                    errores.append(f"{imp['nombre']}: {err}")
                    continue
                try:
                    import io as _io
                    filas = []
                    for sep in (";", ",", "\t"):
                        import csv as _csv
                        reader = _csv.DictReader(_io.StringIO(csv_txt), delimiter=sep)
                        raw = list(reader)
                        if raw and len(raw[0]) > 2:
                            filas = self._dlg_contabilidad._procesar_filas(raw)
                            break
                    if filas:
                        if ip not in datos:
                            datos[ip] = {"nombre_impresora": f"{imp['nombre']} ({ip})", "snapshots": {}}
                        datos[ip]["snapshots"][mes_actual] = {
                            "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "usuarios": filas,
                        }
                        ok += 1
                    else:
                        errores.append(f"{imp['nombre']}: CSV sin datos de usuario")
                except Exception as e:
                    errores.append(f"{imp['nombre']}: {e}")

            guardar_json(CONTABILIDAD_FILE, datos)
            self.cfg["xsa_ultimo_mes"] = mes_actual
            guardar_json(CONFIG_FILE, self.cfg)

            def done():
                self._dlg_contabilidad._datos = datos
                self._dlg_contabilidad._refresh_controls()
                self._dlg_contabilidad._poblar()
                msg = f"✔ Descarga automática XSA completada: {ok} impresora(s)."
                if errores:
                    msg += "\nErrores:\n" + "\n".join(errores)
                self.lbl_scan.configure(text=msg)
            self.after(0, done)
            self._schedule_xsa_autodownload()

        threading.Thread(target=worker, daemon=True).start()

    # ── CRUD ──────────────────────────────────────────────────────────────────
    def _añadir(self):
        d = DialogImpresora(self, self.cfg)
        self.wait_window(d)
        if d.resultado:
            ips = [i["ip"] for i in self.impresoras]
            if d.resultado["ip"] in ips:
                messagebox.showwarning("Duplicado", f"{d.resultado['ip']} ya existe.")
                return
            self.impresoras.append(d.resultado)
            guardar_json(DB_FILE, self.impresoras)
            self._poblar_tabla()
            ip = d.resultado["ip"]
            threading.Thread(target=lambda: self._escanear_one(ip), daemon=True).start()

    def _editar(self):
        if not self.sel_ip:
            messagebox.showinfo("", "Selecciona un dispositivo primero.")
            return
        imp = next((i for i in self.impresoras if i["ip"]==self.sel_ip), None)
        if not imp: return
        d = DialogImpresora(self, self.cfg, imp=imp)
        self.wait_window(d)
        if d.resultado:
            imp.update(d.resultado)
            guardar_json(DB_FILE, self.impresoras)
            self._poblar_tabla()
            self._panel_detalle(self.sel_ip)

    def _eliminar(self):
        if not self.sel_ip:
            messagebox.showinfo("", "Selecciona un dispositivo primero.")
            return
        imp = next((i for i in self.impresoras if i["ip"]==self.sel_ip), None)
        if not messagebox.askyesno("Eliminar", f"¿Eliminar '{imp['nombre']}' ({self.sel_ip})?"):
            return
        self.impresoras = [i for i in self.impresoras if i["ip"]!=self.sel_ip]
        self.cache.pop(self.sel_ip, None)
        guardar_json(DB_FILE, self.impresoras)
        self.sel_ip = None
        self._panel_vacio()
        self._poblar_tabla()

    def _escanear_one(self, ip):
        imp = next((i for i in self.impresoras if i["ip"]==ip), {})
        com = imp.get("comunidad") or self.cfg["comunidad_snmp"]
        cons, err = obtener_consumibles(ip, com, self.cfg["timeout_snmp"])
        info = obtener_info_extra(ip, com, self.cfg["timeout_snmp"]) if not err else {}
        alerts = obtener_alertas(ip, com, self.cfg["timeout_snmp"]) if not err else []
        if err:
            est = "offline"
        else:
            tc = any(c["porcentaje"]<=self.cfg["umbral_critico"] for c in (cons or []))
            ta = any(c["porcentaje"]<=self.cfg["umbral_alerta"]  for c in (cons or []))
            est = "critico" if tc else "alerta" if ta else "ok"
            if cons: guardar_historial(ip, cons)
        self.cache[ip] = {"consumibles":cons,"error":err,"ts":datetime.now(),"estado":est,"info":info,"alerts":alerts}
        self.after(0, self._poblar_tabla)

    # ── ESCANEO RED ───────────────────────────────────────────────────────────
    def _escanear_red(self):
        ips_existentes = [i["ip"] for i in self.impresoras]
        d = DialogEscaneoRed(self, self.cfg, ips_existentes)
        self.wait_window(d)
        if not d.resultado:
            return
        añadidas = 0
        ips_ya = {i["ip"] for i in self.impresoras}
        for imp in d.resultado:
            if imp["ip"] in ips_ya:
                continue
            self.impresoras.append(imp)
            añadidas += 1
        if añadidas:
            guardar_json(DB_FILE, self.impresoras)
            self._poblar_tabla()
            # Escanear consumibles de las recién añadidas en background
            for imp in d.resultado:
                ip = imp["ip"]
                threading.Thread(target=lambda _ip=ip: self._escanear_one(_ip),
                                 daemon=True).start()
            messagebox.showinfo("Añadidas", f"Se añadieron {añadidas} impresoras.")

    # ── CONFIG ────────────────────────────────────────────────────────────────
    def _abrir_config(self):
        def on_save():
            self._schedule_autoref()
            self._poblar_tabla()
        DialogConfig(self, self.cfg, on_save)

    # ── EXPORTAR ──────────────────────────────────────────────────────────────
    def _recopilar_filas(self, ip=None):
        filas = []
        ips = [ip] if ip else [i["ip"] for i in self.impresoras]
        for _ip in ips:
            imp   = next((i for i in self.impresoras if i["ip"]==_ip), {})
            cache = self.cache.get(_ip, {})
            info  = cache.get("info", {})
            estado= cache.get("estado","")
            ts_s  = cache["ts"].strftime("%Y-%m-%d %H:%M:%S") if cache.get("ts") else ""
            pag   = info.get("paginas","—")
            modelo= info.get("modelo","—")
            serial= info.get("serial","—")
            uptime= _formatear_uptime(info["sys_uptime"]) if info.get("sys_uptime") else "—"
            cons  = cache.get("consumibles") or []
            if cons:
                for c in cons:
                    filas.append({
                        "IP": _ip, "Nombre": imp.get("nombre",""),
                        "Ubicacion": imp.get("ubicacion",""),
                        "Modelo": modelo, "Serie": serial,
                        "Paginas": pag, "Uptime": uptime,
                        "Consumible": c["componente"], "Porcentaje": c["porcentaje"],
                        "Estado": estado.upper(), "Timestamp": ts_s
                    })
            else:
                filas.append({
                    "IP": _ip, "Nombre": imp.get("nombre",""),
                    "Ubicacion": imp.get("ubicacion",""),
                    "Modelo": modelo, "Serie": serial,
                    "Paginas": pag, "Uptime": uptime,
                    "Consumible": "—", "Porcentaje": "—",
                    "Estado": estado.upper() or "SIN DATOS", "Timestamp": ts_s
                })
        return filas

    def _exportar_csv(self, ip=None):
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(
            defaultextension=".csv", filetypes=[("CSV","*.csv")],
            initialfile=f"reporte_{datetime.now().strftime('%Y%m%d_%H%M')}.csv")
        if not path: return
        filas = self._recopilar_filas(ip)
        if not filas:
            messagebox.showinfo("", "Sin datos para exportar.")
            return
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=filas[0].keys())
            w.writeheader(); w.writerows(filas)
        messagebox.showinfo("Exportado", f"CSV guardado en:\n{path}")

    def _exportar_excel(self, ip=None):
        from tkinter.filedialog import asksaveasfilename
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Error", "Instala openpyxl:\npip install openpyxl")
            return

        path = asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile=f"reporte_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        if not path: return

        filas = self._recopilar_filas(ip)
        if not filas:
            messagebox.showinfo("", "Sin datos para exportar.")
            return

        wb = openpyxl.Workbook()

        # ── Hoja 1: Resumen por dispositivo ──────────────────────────────────
        ws1 = wb.active
        ws1.title = "Resumen"
        ws1.sheet_view.showGridLines = False
        ws1.freeze_panes = "A2"

        # Colores
        fill_header  = PatternFill("solid", fgColor="1A1D2E")
        fill_ok      = PatternFill("solid", fgColor="1A3A2A")
        fill_alerta  = PatternFill("solid", fgColor="3A2A00")
        fill_critico = PatternFill("solid", fgColor="3A0A0A")
        fill_offline = PatternFill("solid", fgColor="2A2A2A")
        fill_alt     = PatternFill("solid", fgColor="1E2238")

        font_header = Font(name="Segoe UI", bold=True, color="E8EAF0", size=10)
        font_text   = Font(name="Segoe UI", color="E8EAF0", size=10)
        font_ok     = Font(name="Segoe UI", color="2ECC71", bold=True, size=10)
        font_warn   = Font(name="Segoe UI", color="F39C12", bold=True, size=10)
        font_crit   = Font(name="Segoe UI", color="E74C3C", bold=True, size=10)
        font_off    = Font(name="Segoe UI", color="6C7A99", size=10)
        align_c     = Alignment(horizontal="center", vertical="center")
        align_l     = Alignment(horizontal="left",   vertical="center")

        thin = Side(style="thin", color="353860")
        border = Border(bottom=thin)

        # Cabecera resumen
        hdrs1 = ["IP","Nombre","Ubicación","Modelo","Serie","Páginas","Estado","Última lect."]
        ws1.row_dimensions[1].height = 22
        for ci, h in enumerate(hdrs1, 1):
            c = ws1.cell(row=1, column=ci, value=h)
            c.fill = fill_header; c.font = font_header; c.alignment = align_c

        # Datos resumen (una fila por impresora)
        vistos = set()
        fila_num = 2
        for f in filas:
            ip_f = f["IP"]
            if ip_f in vistos: continue
            vistos.add(ip_f)
            estado_f = f["Estado"]
            fill_f = {"OK":fill_ok,"ALERTA":fill_alerta,"CRITICO":fill_critico}.get(estado_f, fill_offline)
            font_e = {"OK":font_ok,"ALERTA":font_warn,"CRITICO":font_crit}.get(estado_f, font_off)
            vals = [ip_f, f["Nombre"], f["Ubicacion"], f["Modelo"], f["Serie"],
                    f["Paginas"], estado_f, f["Timestamp"]]
            ws1.row_dimensions[fila_num].height = 20
            for ci, v in enumerate(vals, 1):
                c = ws1.cell(row=fila_num, column=ci, value=v)
                c.fill = fill_f
                c.font = font_e if ci==7 else font_text
                c.alignment = align_c if ci in (1,6,7,8) else align_l
                c.border = border
            fila_num += 1

        anchos1 = [16,28,20,30,18,12,12,18]
        for ci, w in enumerate(anchos1, 1):
            ws1.column_dimensions[get_column_letter(ci)].width = w

        # ── Hoja 2: Consumibles detalle ───────────────────────────────────────
        ws2 = wb.create_sheet("Consumibles")
        ws2.sheet_view.showGridLines = False
        ws2.freeze_panes = "A2"

        hdrs2 = ["IP","Nombre","Ubicación","Consumible","Porcentaje","Estado","Timestamp"]
        ws2.row_dimensions[1].height = 22
        for ci, h in enumerate(hdrs2, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.fill = fill_header; c.font = font_header; c.alignment = align_c

        alt = False
        for fi, f in enumerate(filas, 2):
            pct = f["Porcentaje"]
            try: pct_int = int(pct)
            except: pct_int = -1

            if pct_int >= 0:
                if pct_int <= self.cfg["umbral_critico"]:
                    fill_f, font_p = fill_critico, font_crit
                elif pct_int <= self.cfg["umbral_alerta"]:
                    fill_f, font_p = fill_alerta, font_warn
                else:
                    fill_f, font_p = (fill_alt if alt else fill_ok), font_ok
            else:
                fill_f, font_p = fill_offline, font_off

            alt = not alt
            vals2 = [f["IP"],f["Nombre"],f["Ubicacion"],f["Consumible"],
                     f"{pct}%" if pct_int>=0 else pct, f["Estado"], f["Timestamp"]]
            ws2.row_dimensions[fi].height = 20
            for ci, v in enumerate(vals2, 1):
                c = ws2.cell(row=fi, column=ci, value=v)
                c.fill = fill_f
                c.font = font_p if ci==5 else font_text
                c.alignment = align_c if ci in (1,5,6,7) else align_l
                c.border = border

        anchos2 = [16,28,20,28,13,12,18]
        for ci, w in enumerate(anchos2, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = w

        # ── Hoja 3: Solo alertas ──────────────────────────────────────────────
        ws3 = wb.create_sheet("Alertas")
        ws3.sheet_view.showGridLines = False
        ws3.freeze_panes = "A2"
        for ci, h in enumerate(hdrs2, 1):
            c = ws3.cell(row=1, column=ci, value=h)
            c.fill = fill_header; c.font = font_header; c.alignment = align_c

        fila_alerta = 2
        for f in filas:
            pct = f["Porcentaje"]
            try: pct_int = int(pct)
            except: pct_int = 101
            if pct_int > self.cfg["umbral_alerta"] and f["Estado"] not in ("OFFLINE","SIN DATOS"):
                continue
            fill_f = fill_critico if pct_int <= self.cfg["umbral_critico"] else fill_alerta
            font_p = font_crit    if pct_int <= self.cfg["umbral_critico"] else font_warn
            vals3 = [f["IP"],f["Nombre"],f["Ubicacion"],f["Consumible"],
                     f"{pct}%" if pct_int<=100 else pct, f["Estado"], f["Timestamp"]]
            ws3.row_dimensions[fila_alerta].height = 20
            for ci, v in enumerate(vals3, 1):
                c = ws3.cell(row=fila_alerta, column=ci, value=v)
                c.fill = fill_f
                c.font = font_p if ci==5 else font_text
                c.alignment = align_c if ci in (1,5,6,7) else align_l
                c.border = border
            fila_alerta += 1

        for ci, w in enumerate(anchos2, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = w

        wb.save(path)
        messagebox.showinfo("Exportado", f"Excel guardado en:\n{path}\n\nHojas: Resumen · Consumibles · Alertas")

    def _exportar_todo(self):
        # Preguntar formato
        from tkinter.simpledialog import askstring
        fmt = askstring("Formato", "Exportar como:\n  1 → Excel (.xlsx)\n  2 → CSV (.csv)", parent=self)
        if fmt == "1":
            self._exportar_excel()
        elif fmt == "2":
            self._exportar_csv()

# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()